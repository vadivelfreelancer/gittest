# -*- coding: utf-8 -*-

from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from io import StringIO,BytesIO
import base64

class expenditure_proposal_report_wizard(models.TransientModel):
    _name='expenditure.proposal.report.wizard'
    _description='Expenditure proposal report'
    
    state_id = fields.Many2one('res.country.state', 'State')
    district_id = fields.Many2one('state.district', 'District')
    cluster_id = fields.Many2one('pappaya.cluster', 'Cluster')
    division_id = fields.Many2one('pappaya.division', 'Division')
    
    @api.multi
    def _get_expenditure_data(self):
        expenditure_data = []
        sno = 0
        for expenditure in self.env['expenditure.proposal'].search([]):
            expenditure_data_dict = {}
            for line in expenditure.expenditure_proposal_line:
                sno += 1
                expenditure_data_dict.update({
                    'sno':sno,
                    'division':expenditure.division_id.name or ' ',
                    'staff_name':line.staff_id.name or ' ',
                    'proposed_amt':line.amount,
                    'approved_amt':line.approved_amount,
                    'bank_name': line.bank_id.name or ' ',
                    'account_number':line.account_number,
                    'cheque_no': ' ',
                    'transaction_id': ' ',
                    'voucher_no':' ',
                    'nap_status':line.state,
                    'nerp_status':' '
                    
                    })
                expenditure_data.append(expenditure_data_dict)
        return expenditure_data
    
    @api.multi
    def action_get_pdf(self):
        if not self._get_expenditure_data():
            raise ValidationError('No records found.')
        return self.env.ref('pappaya_marketing.report_template_expenditure_proposal_pdf').report_action(self)
    
    @api.multi
    def generate_excel_report(self):
        wb= Workbook()
        ws= wb.active
        ws.append(['NAP Expenditure Report'])
        ws.append(['Sno','Division','Ref No','Staff Name','Proposed Amt','Approved Amt','Bank Name','Account No',
                    'Cheque No','Transaction Id','Voucher No','NAP Status','NERP Status'])
        sno = 0
        for expenditure in self.env['expenditure.proposal'].search([]):
            for line in expenditure.expenditure_proposal_line:
                sno += 1
                ws.append([sno,
                        (expenditure.division_id.name if expenditure.division_id else ''),
                        (line.ref_no if line.ref_no else ''),
                        (line.staff_id.name if line.staff_id else ''),
                        (line.amount if line.amount else ''),
                        (line.approved_amount if line.approved_amount else ''),
                        (line.bank_id.name if line.bank_id else ''),
                        (line.account_number if line.account_number else ''),
                        '','','',line.state,''])
                
        """ Alignment & Designs """
        
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=15, bold=True, color='058598')
        header1 = NamedStyle(name="header1", font=ft1)
        ws['A1'].style = header1
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:N1')
        font_bold = Font(bold=True)
        header6 = NamedStyle(name="header6", font=font_bold)
        ws['A2'].style = header6;ws['B2'].style = header6;ws['C2'].style = header6;ws['D2'].style = header6
        ws['E2'].style = header6;ws['F2'].style = header6;ws['G2'].style = header6;ws['H2'].style = header6
        ws['I2'].style = header6;ws['J2'].style = header6;ws['K2'].style = header6;ws['L2'].style = header6
        ws['M2'].style = header6;
        
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['L'].width = 18
         
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data    
    
    @api.multi
    def action_get_excel(self):
        data = base64.encodestring(self.generate_excel_report())
        attach_vals = {
            'name': '%s.xls' % ('Expenditure proposal'),
            'datas': data,
            'datas_fname': '%s.xls' % ('Expenditure Proposal'),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        
        print (doc_id, "doc_iddoc_id")
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }