from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import base64
from io import BytesIO
import os
import openpyxl
from openpyxl.drawing.image import Image
from tempfile import TemporaryFile
import re
from lxml import etree
from odoo.osv.orm import setup_modifiers


class AddressCollection(models.TransientModel):
    _name = "address.collection"
    _rec_name = 'name'

    report_type = fields.Selection([('branch','Branch Report'),('branch_employee','Branch & Employee Report'),('employee','Employee Report')], string='Report Type', default='branch')
    state_ids = fields.Many2many('res.country.state', 'state_address_rel', 'address_id', 'state_id', string='State Name')
    district_ids = fields.Many2many('state.district', 'district_address_rel', 'address_id', 'district_id', string='District Name')
    branch_ids = fields.Many2many('res.company','branch_address_rel','address_id','branch_id', string= 'Branch Name')
    payroll_branch_ids = fields.Many2many('pappaya.payroll.branch', 'payroll_branch_address_rel', 'address_id', 'payroll_branch_id', string='Payroll Branch Name')
    employee_ids = fields.Many2many('hr.employee', 'employee_address_rel', 'address_id', 'employee_id', string='Employee Name')
    course_ids = fields.Many2many('pappaya.lead.course', 'course_address_rel', 'address_id', 'course_id', string='Lead Class')
    is_classwise = fields.Boolean(string='Show Classwise', default=False)
    branch_line_ids = fields.One2many('address.collection.branch', 'address_branch_id', string='Branch Report')
    branch_class_line_ids = fields.One2many('address.collection.branch', 'address_branch_id', string='Branch Classwise Report')
    is_branch = fields.Boolean(string='Is Branch?', default=False)
    is_branch_class = fields.Boolean(string='Is Branch Class', default=False)
    branch_emp_line_ids = fields.One2many('address.collection.branch.employee', 'address_branch_employee_id', string='Branch and Employee Report')
    branch_emp_class_line_ids = fields.One2many('address.collection.branch.employee', 'address_branch_employee_id', string='Branch and Employee Classwise Report')
    is_branch_emp = fields.Boolean(string='Is Branch & Employee?', default=False)
    is_branch_emp_class = fields.Boolean(string='Is Branch & Employee Class?', default=False)
    employee_line_ids = fields.One2many('address.collection.employee', 'address_employee_id', string='Employee Report')
    employee_class_line_ids = fields.One2many('address.collection.employee', 'address_employee_id', string='Employee Classwise Report')
    is_employee = fields.Boolean(string='Is Employee?', default=False)
    is_employee_class = fields.Boolean(string='Is Employee Class?', default=False)
    name = fields.Char(string='Display Name', compute='compute_name')

    @api.model
    def fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(AddressCollection, self).fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar,submenu=submenu)
        for rec in self.browse(self._context.get('active_ids')):
            course_list = ['course1', 'course2', 'course3', 'course4', 'course5', 'course6', 'course7', 'course8',
                           'course9', 'course10', 'course11', 'course12', 'course13', 'course14', 'course15']
            doc_list = ''
            if rec.report_type == 'branch':
                doc_list = 'branch_class_line_ids'
            elif rec.report_type == 'branch_employee':
                doc_list = 'branch_emp_class_line_ids'
            elif rec.report_type == 'employee':
                doc_list = 'employee_class_line_ids'
            if view_type == 'form':
                doc = etree.XML(res['fields'][doc_list]['views']['tree']['arch'])
                if rec.is_classwise == True or not rec.course_ids:
                    for obj in self.env['pappaya.lead.course'].search([], order='name asc'):
                        for c_value in course_list:
                            course_name = obj.name
                            nodes = doc.xpath("//field[@name='" + c_value + "']")
                            for node in nodes:
                                node.set('string', str(course_name))
                            course_list.remove(c_value)
                            break
                    for c_val in course_list:
                        nodes = doc.xpath("//field[@name='" + c_val + "']")
                        for node in nodes:
                            node.set('invisible', 'True')
                            doc.remove(node)
                            setup_modifiers(node, res['fields'][doc_list]['views']['tree']['fields'][c_val])
                    res['fields'][doc_list]['views']['tree']['arch'] = etree.tostring(doc)
                elif rec.course_ids:
                    for obj in rec.course_ids:
                        for c_value in course_list:
                            course_name = obj.name
                            nodes = doc.xpath("//field[@name='" + c_value + "']")
                            for node in nodes:
                                node.set('string', str(course_name))
                            course_list.remove(c_value)
                            break
                    for c_val in course_list:
                        nodes = doc.xpath("//field[@name='" + c_val + "']")
                        for node in nodes:
                            node.set('invisible', 'True')
                            doc.remove(node)
                            setup_modifiers(node, res['fields'][doc_list]['views']['tree']['fields'][c_val])
                    res['fields'][doc_list]['views']['tree']['arch'] = etree.tostring(doc)
        return res

    @api.depends('report_type')
    def compute_name(self):
        report_dict = {'branch':'Branch Report','branch_employee':'Branch & Employee Report','employee':'Employee Report'}
        self.name = report_dict[self.report_type]

    @api.onchange('report_type')
    def onchange_report_type(self):
        if self.report_type:
            self.state_ids = []
            self.district_ids = []
            self.branch_ids = []
            self.payroll_branch_ids = []
            self.course_ids = []
            self.is_classwise = False
        if self.report_type == 'branch':
            self.is_branch = True
            self.is_branch_emp = False
            self.is_employee = False
        elif self.report_type == 'branch_employee':
            self.is_branch_emp = True
            self.is_branch = False
            self.is_employee = False
        elif self.report_type == 'employee':
            self.is_employee = True
            self.is_branch = False
            self.is_branch_emp = False

    @api.onchange('state_ids')
    def onchange_state_ids(self):
        if self.state_ids:
            self.district_ids = []
            self.branch_ids = []
            self.payroll_branch_ids = []
            self.employee_ids = []
            self.course_ids = []
            self.is_classwise = False
            return {'domain': {'district_ids': [('state_id','in',self.state_ids.ids)], 'payroll_branch_ids': [('state_id','in',self.state_ids.ids)]}}

    @api.onchange('district_ids')
    def onchange_district_ids(self):
        if self.district_ids:
            self.branch_ids = []
            self.payroll_branch_ids = []
            self.employee_ids = []
            self.course_ids = []
            self.is_classwise = False
            return {'domain': {'branch_ids': [('id', '!=', 1),('type','=','branch'),('state_id','in',self.state_ids.ids),('state_district_id','in',self.district_ids.ids)]}}

    @api.onchange('course_ids')
    def onchange_course_ids(self):
        if self.course_ids and self.report_type == 'branch':
            self.is_branch = False
        elif self.course_ids and self.report_type == 'branch_employee':
            self.is_branch_emp = False
        elif self.course_ids and self.report_type == 'employee':
            self.is_employee = False

    @api.onchange('payroll_branch_ids')
    def onchange_payroll_branch_ids(self):
        if self.payroll_branch_ids:
            self.employee_ids = []
            return {'domain': {'employee_ids': [('id','!=',1),('payroll_branch_id','in',self.payroll_branch_ids.ids)]}}

    @api.onchange('is_classwise')
    def onchange_classwise(self):
        if self.is_classwise == True:
            self.course_ids = []
        if self.is_classwise == True and self.report_type == 'branch':
            self.is_branch = False
        elif self.is_classwise == True and self.report_type == 'branch_employee':
            self.is_branch_emp = False
        elif self.is_classwise == True and self.report_type == 'employee':
            self.is_employee = False
        if self.is_classwise == False and self.report_type == 'branch':
            self.is_branch = True
            self.is_branch_class = False
        elif self.is_classwise == False and self.report_type == 'branch_employee':
            self.is_branch_emp = True
            self.is_branch_emp_class = False
        elif self.is_classwise == False and self.report_type == 'employee':
            self.is_employee = True
            self.is_employee_class = False

    @api.multi
    def get_course(self):
        course_list = []
        if self.is_classwise or not self.course_ids:
            for course in self.env['pappaya.lead.course'].search([], order='name asc'):
                vals = {}
                vals['course_name'] = course.name
                course_list.append(vals)
        if self.course_ids:
            for course in self.env['pappaya.lead.course'].search([('id', 'in', self.course_ids.ids)], order='id asc'):
                vals = {}
                vals['course_name'] = course.name
                course_list.append(vals)
        return course_list

    # Branch Report

    @api.multi
    def get_branch_data(self):
        domain, data_list, branch_dict = [], [],{}
        if self.state_ids:
            domain.append(('branch_id.state_id', 'in', self.state_ids.ids))
        if self.district_ids:
            domain.append(('branch_id.state_district_id', 'in', self.district_ids.ids))
        if self.branch_ids:
            domain.append(('branch_id', 'in', self.branch_ids.ids))
        if self.course_ids:
            domain.append(('studying_course_id', 'in', self.course_ids.ids))
        addr_obj = self.env['pappaya.lead.stud.address'].search(domain)
        for obj in addr_obj:
            if obj.branch_id.name not in branch_dict:
                branch_dict.update({obj.branch_id.name: [obj]})
            else:
                branch_dict[obj.branch_id.name].append(obj)
        s_no = 0
        for branch_key in branch_dict:
            s_no += 1
            vals = {}
            vals['s_no'] = s_no
            vals['branch_name'] = branch_key
            vals['course'] = []
            crse_list, t_count = [], 0
            if self.is_classwise == True or not self.course_ids:
                cl_list = []
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    cl_list.append(cl.name)
                for course_key in cl_list:
                    cl_count = 0
                    for addr in self.env['pappaya.lead.stud.address'].search([('branch_id.name','=',branch_key),('studying_course_id.name','=',course_key)]):
                        if addr:
                            cl_count += 1
                    cl_vals = {}
                    if cl_count > 0.0:
                        crse_list.append(cl_count or 0)
                        cl_vals['c_count'] = cl_count or 0
                    else:
                        crse_list.append(0)
                        cl_vals['c_count'] = 0
                    vals['course'].append(cl_vals)
                    t_count += cl_count
                vals['course_count'] = crse_list
            elif self.course_ids:
                for cl_key in self.course_ids:
                    cl_count = 0
                    for addr in self.env['pappaya.lead.stud.address'].search([('branch_id.name', '=', branch_key),('studying_course_id.name','=',cl_key.name)]):
                        if addr:
                            cl_count += 1
                    cl_vals = {}
                    if cl_count > 0.0:
                        crse_list.append(cl_count or 0)
                        cl_vals['c_count'] = cl_count or 0
                    else:
                        crse_list.append(0)
                        cl_vals['c_count'] = 0
                    vals['course'].append(cl_vals)
                    t_count += cl_count
                vals['course_count'] = crse_list
            vals['address_count'] = t_count
            data_list.append(vals)
        return data_list

    @api.multi
    def get_branch(self):
        data_list, domain, branch_dict = [], [], {}
        if self.state_ids:
            domain.append(('branch_id.state_id', 'in', self.state_ids.ids))
        if self.district_ids:
            domain.append(('branch_id.state_district_id', 'in', self.district_ids.ids))
        if self.branch_ids:
            domain.append(('branch_id', 'in', self.branch_ids.ids))
        if self.course_ids:
            domain.append(('studying_course_id', 'in', self.course_ids.ids))
        addr_obj = self.env['pappaya.lead.stud.address'].search(domain)
        for obj in addr_obj:
            if obj.branch_id.name not in branch_dict:
                branch_dict.update({obj.branch_id.id: [obj]})
            else:
                branch_dict[obj.branch_id.id].append(obj)
        s_no = 0
        for branch_key in branch_dict:
            s_no += 1
            vals = {}
            vals['s_no'] = s_no
            vals['branch_id'] = branch_key
            vals['course1'], vals['course2'], vals['course3'], vals['course4'], vals['course5'], vals['course6'],vals['course7'], vals['course8'] = 0, 0, 0, 0, 0, 0, 0, 0
            vals['course9'], vals['course10'], vals['course11'], vals['course12'], vals['course13'], vals['course14'], vals['course15'] = 0, 0, 0, 0, 0, 0, 0
            addr_count, cnt = 0, 0
            if self.is_classwise == True or not self.course_ids:
                cl_list = []
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    cl_list.append(cl.name)
                for cl_key in cl_list:
                    cnt += 1
                    cl_count = 0
                    for addr in self.env['pappaya.lead.stud.address'].search([('branch_id','=',branch_key),('studying_course_id.name','=',cl_key)]):
                        num_list = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15]
                        for num in num_list:
                            if addr and cnt == num:
                                vals['course'+str(num)] += 1
                                cl_count += 1
                    addr_count += cl_count
            elif self.course_ids:
                for course_key in self.course_ids:
                    cnt += 1
                    cl_count = 0
                    for addr in self.env['pappaya.lead.stud.address'].search([('branch_id', '=', branch_key), ('studying_course_id.name', '=', course_key.name)]):
                        num_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
                        for num in num_list:
                            if addr and cnt == num:
                                vals['course' + str(num)] += 1
                                cl_count += 1
                    addr_count += cl_count
            vals['address_count'] = addr_count
            vals['address_branch_id'] = self.id
            vals['report_type'] = self.report_type
            data_list.append((0, 0, vals))
        return data_list

    @api.multi
    def action_branch_address(self):
        data = self.get_branch()
        if len(self.branch_line_ids) == 0:
            self.write({'branch_line_ids': data})
        else:
            self.branch_line_ids.unlink()
            self.write({'branch_line_ids': data})
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.ids[0],
            'res_model': 'address.collection',
            'target': 'new',
        }

    @api.multi
    def action_branch_class_address(self):
        self.is_branch_class = True
        data = self.get_branch()
        if len(self.branch_class_line_ids) == 0:
            self.write({'branch_class_line_ids': data})
        else:
            self.branch_class_line_ids.unlink()
            self.write({'branch_class_line_ids': data})
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.ids[0],
            'res_model': 'address.collection',
            'target': 'new',
        }

    # Branch and Employee Report

    @api.multi
    def get_branch_emp_data(self):
        domain, data_list, branch_dict, emp_dict = [], [], {}, {}
        if self.state_ids:
            domain.append(('branch_id.state_id', 'in', self.state_ids.ids))
        if self.district_ids:
            domain.append(('branch_id.state_district_id', 'in', self.district_ids.ids))
        if self.branch_ids:
            domain.append(('branch_id', 'in', self.branch_ids.ids))
        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))
        if self.course_ids:
            domain.append(('studying_course_id', 'in', self.course_ids.ids))
        addr_obj = self.env['pappaya.lead.stud.address'].search(domain)
        s_no = 1
        for obj in addr_obj:
            if obj.branch_id.name not in branch_dict:
                branch_dict.update({obj.branch_id.name: [obj]})
            else:
                branch_dict[obj.branch_id.name].append(obj)
            if obj.employee_id.name not in emp_dict:
                emp_dict.update({obj.employee_id.name: [obj]})
            else:
                emp_dict[obj.employee_id.name].append(obj)
        for branchkey in branch_dict:
            for empkey in emp_dict:
                count, emp_name, emp_id, emp_mob = 0, '', '', ''
                for stud_addr in self.env['pappaya.lead.stud.address'].search([('branch_id.name','=',branchkey),('employee_id.name','=',empkey)]):
                    emp_name =  stud_addr.employee_id.name
                    emp_id = stud_addr.employee_id.emp_id
                    emp_mob = stud_addr.employee_id.work_mobile
                vals = {}
                vals['s_no'] = s_no
                vals['employee_name'] = emp_name
                vals['employee_id'] = emp_id
                vals['mobile_no'] = emp_mob
                vals['branch_name'] = branchkey
                vals['course'] = []
                crse_list, t_count = [], 0
                if self.is_classwise == True or not self.course_ids:
                    cl_list = []
                    for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                        cl_list.append(cl.name)
                    for course_key in cl_list:
                        cl_count = 0
                        for addr in self.env['pappaya.lead.stud.address'].search([('branch_id.name', '=', branchkey),('employee_id.name', '=', empkey),('studying_course_id.name', '=', course_key)]):
                            if addr:
                                cl_count += 1
                        cl_vals = {}
                        if cl_count > 0.0:
                            crse_list.append(cl_count or 0)
                            cl_vals['c_count'] = cl_count or 0
                        else:
                            crse_list.append(0)
                            cl_vals['c_count'] = 0
                        vals['course'].append(cl_vals)
                        t_count += cl_count
                    vals['course_count'] = crse_list
                elif self.course_ids:
                    for cl_key in self.course_ids:
                        cl_count = 0
                        for addr in self.env['pappaya.lead.stud.address'].search([('branch_id.name', '=', branchkey),('employee_id.name', '=', empkey),('studying_course_id.name', '=', cl_key.name)]):
                            if addr:
                                cl_count += 1
                        cl_vals = {}
                        if cl_count > 0.0:
                            crse_list.append(cl_count or 0)
                            cl_vals['c_count'] = cl_count or 0
                        else:
                            crse_list.append(0)
                            cl_vals['c_count'] = 0
                        vals['course'].append(cl_vals)
                        t_count += cl_count
                    vals['course_count'] = crse_list
                vals['address_count'] = t_count
                count_list_set =  list(set(vals['course_count']))
                if len(count_list_set) == 1 and count_list_set[0] == 0:
                    pass
                else:
                    data_list.append(vals)
                    s_no += 1
        return data_list

    @api.multi
    def get_branch_emp(self):
        data_list, domain, branch_dict, emp_dict = [], [], {}, {}
        if self.state_ids:
            domain.append(('branch_id.state_id', 'in', self.state_ids.ids))
        if self.district_ids:
            domain.append(('branch_id.state_district_id', 'in', self.district_ids.ids))
        if self.branch_ids:
            domain.append(('branch_id', 'in', self.branch_ids.ids))
        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))
        if self.course_ids:
            domain.append(('studying_course_id', 'in', self.course_ids.ids))
        addr_obj = self.env['pappaya.lead.stud.address'].search(domain)
        for obj in addr_obj:
            if obj.branch_id.id not in branch_dict:
                branch_dict.update({obj.branch_id.id: [obj]})
            else:
                branch_dict[obj.branch_id.id].append(obj)
            if obj.employee_id.id not in emp_dict:
                emp_dict.update({obj.employee_id.id: [obj]})
            else:
                emp_dict[obj.employee_id.id].append(obj)
        s_no = 1
        for branchkey in branch_dict:
            for empkey in emp_dict:
                employee_id, emp_id, emp_mob, branch_id = '', '', '', ''
                for stud_addr in self.env['pappaya.lead.stud.address'].search([('branch_id', '=', branchkey),('employee_id', '=', empkey)]):
                    employee_id = stud_addr.employee_id.id
                    emp_id = stud_addr.employee_id.emp_id
                    emp_mob = stud_addr.employee_id.work_mobile
                    branch_id = stud_addr.branch_id.id
                vals = {}
                vals['s_no'] = s_no
                vals['emp_id'] = emp_id
                vals['employee_id'] = employee_id
                vals['emp_mobile'] = emp_mob
                vals['branch_id'] = branch_id
                vals['course1'], vals['course2'], vals['course3'], vals['course4'], vals['course5'], vals['course6'], vals['course7'], vals['course8'] = 0, 0, 0, 0, 0, 0, 0, 0
                vals['course9'], vals['course10'], vals['course11'], vals['course12'], vals['course13'], vals['course14'], vals['course15'] = 0, 0, 0, 0, 0, 0, 0
                crse_list, cnt, addr_count = [], 0, 0
                if self.is_classwise == True or not self.course_ids:
                    cl_list = []
                    for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                        cl_list.append(cl.name)
                    for cl_key in cl_list:
                        cnt += 1
                        cl_count = 0
                        for addr in self.env['pappaya.lead.stud.address'].search([('branch_id','=',branchkey),('employee_id','=',empkey),('studying_course_id.name', '=', cl_key)]):
                            num_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
                            for num in num_list:
                                if addr and cnt == num:
                                    vals['course' + str(num)] += 1
                                    cl_count += 1
                        addr_count += cl_count
                        cl_vals = {}
                        if cl_count > 0.0:
                            crse_list.append(cl_count or 0)
                            cl_vals['c_count'] = cl_count or 0
                        else:
                            crse_list.append(0)
                            cl_vals['c_count'] = 0
                    vals['course_count'] = crse_list
                elif self.course_ids:
                    for course_key in self.course_ids:
                        cnt += 1
                        cl_count = 0
                        for addr in self.env['pappaya.lead.stud.address'].search([('branch_id','=',branchkey),('employee_id','=',empkey),('studying_course_id.name','=',course_key.name)]):
                            num_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
                            for num in num_list:
                                if addr and cnt == num:
                                    vals['course' + str(num)] += 1
                                    cl_count += 1
                        addr_count += cl_count
                        cl_vals = {}
                        if cl_count > 0.0:
                            crse_list.append(cl_count or 0)
                            cl_vals['c_count'] = cl_count or 0
                        else:
                            crse_list.append(0)
                            cl_vals['c_count'] = 0
                    vals['course_count'] = crse_list
                vals['address_count'] = addr_count
                vals['address_branch_employee_id'] = self.id
                vals['report_type'] = self.report_type
                count_list_set = list(set(vals['course_count']))
                if len(count_list_set) == 1 and count_list_set[0] == 0:
                    pass
                else:
                    data_list.append((0, 0, vals))
                    s_no += 1
        return data_list

    @api.multi
    def action_branch_emp_address(self):
        data = self.get_branch_emp()
        if len(self.branch_emp_line_ids) == 0:
            self.write({'branch_emp_line_ids': data})
        else:
            self.branch_emp_line_ids.unlink()
            self.write({'branch_emp_line_ids': data})
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.ids[0],
            'res_model': 'address.collection',
            'target': 'new',
        }

    @api.multi
    def action_branch_emp_class_address(self):
        self.is_branch_emp_class = True
        data = self.get_branch_emp()
        if len(self.branch_emp_class_line_ids) == 0:
            self.write({'branch_emp_class_line_ids': data})
        else:
            self.branch_emp_class_line_ids.unlink()
            self.write({'branch_emp_class_line_ids': data})
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.ids[0],
            'res_model': 'address.collection',
            'target': 'new',
        }

    # Employee Report

    @api.multi
    def get_employee_data(self):
        domain, data_list, emp_dict = [], [], {}
        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))
        if self.course_ids:
            domain.append(('studying_course_id', 'in', self.course_ids.ids))
        addr_obj = self.env['pappaya.lead.stud.address'].search(domain)
        for obj in addr_obj:
            if obj.employee_id.name not in emp_dict:
                emp_dict.update({obj.employee_id.name: [obj]})
            else:
                emp_dict[obj.employee_id.name].append(obj)
        s_no = 0
        for emp_key in emp_dict:
            s_no += 1
            emp_id, emp_mob = '', ''
            for stud_addr in self.env['pappaya.lead.stud.address'].search([('employee_id.name','=',emp_key)]):
                emp_id = stud_addr.employee_id.emp_id
                emp_mob = stud_addr.employee_id.work_mobile
            vals = {}
            vals['s_no'] = s_no
            vals['employee_id'] = emp_id
            vals['employee_name'] = emp_key
            vals['mobile_no'] = emp_mob
            vals['course'] = []
            crse_list, t_count = [], 0
            if self.is_classwise == True or not self.course_ids:
                cl_list = []
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    cl_list.append(cl.name)
                for cl_key in cl_list:
                    cl_count = 0
                    for addr in self.env['pappaya.lead.stud.address'].search([('employee_id.name','=',emp_key),('studying_course_id.name','=',cl_key)]):
                        if addr:
                            cl_count += 1
                    cl_vals = {}
                    if cl_count > 0.0:
                        crse_list.append(cl_count or 0)
                        cl_vals['c_count'] = cl_count or 0
                    else:
                        crse_list.append(0)
                        cl_vals['c_count'] = 0
                    vals['course'].append(cl_vals)
                    t_count += cl_count
                vals['course_count'] = crse_list
            elif self.course_ids:
                for cl_key in self.course_ids:
                    cl_count = 0
                    for addr in self.env['pappaya.lead.stud.address'].search([('employee_id.name','=',emp_key),('studying_course_id.name','=',cl_key.name)]):
                        if addr:
                            cl_count += 1
                    cl_vals = {}
                    if cl_count > 0.0:
                        crse_list.append(cl_count or 0)
                        cl_vals['c_count'] = cl_count or 0
                    else:
                        crse_list.append(0)
                        cl_vals['c_count'] = 0
                    vals['course'].append(cl_vals)
                    t_count += cl_count
                vals['course_count'] = crse_list
            vals['address_count'] = t_count
            data_list.append(vals)
        return data_list

    @api.multi
    def get_employee(self):
        domain, data_list, emp_dict = [], [], {}
        if self.employee_ids:
            domain.append(('employee_id', 'in', self.employee_ids.ids))
        if self.course_ids:
            domain.append(('studying_course_id', 'in', self.course_ids.ids))
        addr_obj = self.env['pappaya.lead.stud.address'].search(domain)
        for obj in addr_obj:
            if obj.employee_id.name not in emp_dict:
                emp_dict.update({obj.employee_id.id: [obj]})
            else:
                emp_dict[obj.employee_id.id].append(obj)
        s_no = 0
        for emp_key in emp_dict:
            s_no += 1
            emp_id, employee_id, emp_mob = '', '', ''
            for stud_addr in self.env['pappaya.lead.stud.address'].search([('employee_id','=',emp_key)]):
                emp_id = stud_addr.employee_id.emp_id
                employee_id = stud_addr.employee_id.id
                emp_mob = stud_addr.employee_id.work_mobile
            vals = {}
            vals['s_no'] = s_no
            vals['emp_id'] = emp_id
            vals['employee_id'] = employee_id
            vals['emp_mobile'] = emp_mob
            vals['course1'], vals['course2'], vals['course3'], vals['course4'], vals['course5'], vals['course6'], vals['course7'], vals['course8'] = 0, 0, 0, 0, 0, 0, 0, 0
            vals['course9'], vals['course10'], vals['course11'], vals['course12'], vals['course13'], vals['course14'], vals['course15'] = 0, 0, 0, 0, 0, 0, 0
            cnt, addr_count = 0, 0
            if self.is_classwise == True or not self.course_ids:
                cl_list = []
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    cl_list.append(cl.name)
                for cl_key in cl_list:
                    cnt += 1
                    cl_count = 0
                    for addr in self.env['pappaya.lead.stud.address'].search([('employee_id', '=', emp_key),('studying_course_id.name', '=', cl_key)]):
                        num_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
                        for num in num_list:
                            if addr and cnt == num:
                                vals['course' + str(num)] += 1
                                cl_count += 1
                    addr_count += cl_count
            elif self.course_ids:
                for course_key in self.course_ids:
                    cnt += 1
                    cl_count = 0
                    for addr in self.env['pappaya.lead.stud.address'].search([('employee_id', '=', emp_key),('studying_course_id.name', '=', course_key.name)]):
                        num_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
                        for num in num_list:
                            if addr and cnt == num:
                                vals['course' + str(num)] += 1
                                cl_count += 1
                    addr_count += cl_count
            vals['address_count'] = addr_count
            vals['address_employee_id'] = self.id
            vals['report_type'] = self.report_type
            data_list.append((0, 0, vals))
        return data_list

    @api.multi
    def action_employee_address(self):
        data = self.get_employee()
        if len(self.employee_line_ids) == 0:
            self.write({'employee_line_ids': data})
        else:
            self.employee_line_ids.unlink()
            self.write({'employee_line_ids': data})
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.ids[0],
            'res_model': 'address.collection',
            'target': 'new',
        }

    @api.multi
    def action_employee_class_address(self):
        self.is_employee_class = True
        data = self.get_employee()
        if len(self.employee_class_line_ids) == 0:
            self.write({'employee_class_line_ids': data})
        else:
            self.employee_class_line_ids.unlink()
            self.write({'employee_class_line_ids': data})
        return {
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': self.ids[0],
            'res_model': 'address.collection',
            'target': 'new',
        }

    @api.multi
    def generate_xl(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Address Collection"
        ws.append(['','',''+(self.env.user.company_id.name if self.env.user.company_id.name else '')])
        ws.append(['','',''+(self.env.user.company_id.street if self.env.user.company_id.street else '')])
        ws.append(['','',''+(self.env.user.company_id.street2 if self.env.user.company_id.street2 else '') + ', ' +
                   (self.env.user.company_id.city if self.env.user.company_id.city else '') + ', ' +
                   (self.env.user.company_id.state_id.name if self.env.user.company_id.state_id.name else '') + '-' +
                   (self.env.user.company_id.zip if self.env.user.company_id.zip else '')])
        ws.append(['','','Phone: ' + (self.env.user.company_id.phone if self.env.user.company_id.phone else '') + ', ' +
                   'Email: ' + (self.env.user.company_id.email if self.env.user.company_id.email else '')])
        ws.append(['','',''+(self.env.user.company_id.website if self.env.user.company_id.website else '')])
        ws.append([])
        c_count = 0

        # Branch Report
        if self.report_type == 'branch':
            ws.append(['BRANCH REPORT'])
            ws.append([])
            # Heading Name
            course_list = []
            if self.is_classwise == True and not self.course_ids:
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    course_list.append(cl.name)
                ws.append(['S.No', 'BRANCH NAME'] + course_list + ['ADDRESS COUNT'])
            if self.is_classwise == False and not self.course_ids:
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    course_list.append(cl.name)
                ws.append(['S.No', 'BRANCH NAME', 'ADDRESS COUNT'])
            if self.course_ids:
                for co in self.env['pappaya.lead.course'].search([('id', 'in', self.course_ids.ids)], order='id asc'):
                    course_list.append(co.name)
                ws.append(['S.No', 'BRANCH NAME'] + course_list + ['ADDRESS COUNT'])
            c_count += len(course_list)
            t_count = 9
            # Fetch data
            for b_list in self.get_branch_data():
                if self.is_classwise == True and not self.course_ids:
                    ws.append([b_list['s_no'], b_list['branch_name']] + b_list['course_count'] + [b_list['address_count']])
                    ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
                    ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="left")
                    al_len = 1 + c_count
                    col = 0
                    alphabets = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S','T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1
                if self.is_classwise == False and not self.course_ids:
                    ws.append([b_list['s_no'], b_list['branch_name'], b_list['address_count']])
                    ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
                    ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="left")
                    al_len = 1
                    col = 0
                    alphabets = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S','T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1
                if self.course_ids:
                    ws.append([b_list['s_no'], b_list['branch_name']] + b_list['course_count'] + [b_list['address_count']])
                    ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
                    ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="left")
                    al_len = 1 + c_count
                    col = 0
                    alphabets = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S','T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1

        # Branch and Employee Report
        if self.report_type == 'branch_employee':
            ws.append(['BRANCH & EMPLOYEE REPORT'])
            ws.append([])
            # Heading Name
            course_list = []
            if self.is_classwise == True and not self.course_ids:
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    course_list.append(cl.name)
                ws.append(['S.No', 'BRANCH NAME'] + course_list + ['ADDRESS COUNT', 'EMPLOYEE ID', 'EMPLOYEE NAME', 'MOBILE NUMBER'])
            if self.is_classwise == False and not self.course_ids:
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    course_list.append(cl.name)
                ws.append(['S.No', 'BRANCH NAME', 'ADDRESS COUNT', 'EMPLOYEE ID', 'EMPLOYEE NAME', 'MOBILE NUMBER'])
            if self.course_ids:
                for co in self.env['pappaya.lead.course'].search([('id', 'in', self.course_ids.ids)],order='id asc'):
                    course_list.append(co.name)
                    ws.append(['S.No', 'BRANCH NAME'] + course_list + ['ADDRESS COUNT', 'EMPLOYEE ID', 'EMPLOYEE NAME','MOBILE NUMBER'])
            c_count += len(course_list)
            t_count = 9
            # Fetch data
            for be_list in self.get_branch_emp_data():
                if self.is_classwise == True and not self.course_ids:
                    ws.append([be_list['s_no'], be_list['branch_name']] + be_list['course_count'] + [be_list['address_count'], be_list['employee_id'],be_list['employee_name'], be_list['mobile_no']])
                    ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
                    ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="left")
                    al_len = 4 + c_count
                    col = 0
                    alphabets = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S','T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1
                if self.is_classwise == False and not self.course_ids:
                    ws.append([be_list['s_no'], be_list['branch_name'], be_list['address_count'], be_list['employee_id'],be_list['employee_name'], be_list['mobile_no']])
                    ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
                    ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="left")
                    al_len = 4
                    col = 0
                    alphabets = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S','T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1
                if self.course_ids:
                    ws.append([be_list['s_no'], be_list['branch_name'], be_list['address_count'], be_list['employee_id'], be_list['employee_name'], be_list['mobile_no']])
                    ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
                    ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="left")
                    al_len = 4 + c_count
                    col = 0
                    alphabets = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S','T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1

        # Employee Report
        if self.report_type == 'employee':
            ws.append(['EMPLOYEE REPORT'])
            ws.append([])
            # Heading Name
            course_list = []
            if self.is_classwise == True and not self.course_ids:
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    course_list.append(cl.name)
                ws.append(['S.No', 'EMPLOYEE ID', 'EMPLOYEE NAME', 'MOBILE NUMBER'] + course_list + ['ADDRESS COUNT'])
            if self.is_classwise == False and not self.course_ids:
                for cl in self.env['pappaya.lead.course'].search([], order='name asc'):
                    course_list.append(cl.name)
                ws.append(['S.No', 'EMPLOYEE ID', 'EMPLOYEE NAME', 'MOBILE NUMBER', 'ADDRESS COUNT'])
            if self.course_ids:
                for co in self.env['pappaya.lead.course'].search([('id', 'in', self.course_ids.ids)],order='id asc'):
                    course_list.append(co.name)
                ws.append(['S.No', 'EMPLOYEE ID', 'EMPLOYEE NAME', 'MOBILE NUMBER'] + course_list + ['ADDRESS COUNT'])
            c_count += len(course_list)
            t_count = 9
            # Fetch data
            for e_list in self.get_employee_data():
                if self.is_classwise == True and not self.course_ids:
                    ws.append([e_list['s_no'], e_list['employee_id'], e_list['employee_name'], e_list['mobile_no']] + e_list['course_count'] + [e_list['address_count']])
                    al_len = 5 + c_count
                    col = 0
                    alphabets = ['A','B','C','D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R','S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1
                if self.is_classwise == False and not self.course_ids:
                    ws.append([e_list['s_no'], e_list['employee_id'], e_list['employee_name'], e_list['mobile_no'], e_list['address_count']])
                    al_len = 5
                    col = 0
                    alphabets = ['A','B','C','D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R','S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1
                if self.course_ids:
                    ws.append([e_list['s_no'], e_list['employee_id'], e_list['employee_name'], e_list['mobile_no']] + e_list['course_count'] + [e_list['address_count']])
                    al_len = 5 + c_count
                    col = 0
                    alphabets = ['A','B','C','D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R','S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
                    for al in alphabets:
                        col += 1
                        if col <= al_len:
                            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
                            ws[al + str(t_count + 1)].number_format = '0'
                    t_count += 1

        # Company Style
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=15, bold=True)
        header1 = NamedStyle(name="header1", font=ft1)
        ws['C1'].style = header1
        ws['C1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('C1:F1')
        ft2 = Font(size=10)
        header2 = NamedStyle(name="header2", font=ft2)
        ws['C2'].style = header2
        ws['C2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('C2:F2')
        ws['C3'].style = header2
        ws['C3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('C3:F3')
        ws['C4'].style = header2
        ws['C4'].alignment = Alignment(horizontal="center")
        ws.merge_cells('C4:F4')
        ws['C5'].style = header2
        ws['C5'].alignment = Alignment(horizontal="center")
        ws.merge_cells('C5:F5')
        ws['A6'].style = header2
        ws['A6'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A6:F6')
        # Report name style
        ft4 = Font(size=12, bold=True)
        header3 = NamedStyle(name="header3", font=ft4)
        ws['A7'].style = header3
        ws['A7'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A7:F7')
        ws['A8'].style = header3
        ws['A8'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A8:F8')
        # Column/Border/Width styles
        ft5 = Font(size=11, bold=True)
        header4 = NamedStyle(name="header4", font=ft5)
        col, cnt = 0, 0
        if self.report_type == 'branch' and  (self.is_classwise == True or self.course_ids):
            cnt += 2 + c_count
        elif self.report_type == 'branch_employee' and  (self.is_classwise == True or self.course_ids):
            cnt += 5 + c_count
        elif self.report_type == 'employee' and  (self.is_classwise == True or self.course_ids):
            cnt += 4 + c_count
        elif self.report_type == 'branch' and not self.course_ids:
            cnt += 2
        elif self.report_type == 'branch_employee' and not self.course_ids:
            cnt += 5
        elif self.report_type == 'employee' and not self.course_ids:
            cnt += 4
        alphabets = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V','W', 'X', 'Y', 'Z']
        for al in alphabets:
            col += 1
            if col <= cnt:
                # Columns styles
                ws['A9'].style = header4
                ws['A9'].alignment = Alignment(horizontal="center")
                ws[al + str(9)].style = header4
                ws[al + str(9)].alignment = Alignment(horizontal="center")
                # Border Style
                thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
                ws.cell(row=9, column=1).border = thin_border
                ws.cell(row=9, column=(col + 1)).border = thin_border
                # Width style
                ws.column_dimensions['A'].width = 15
                # ws.column_dimensions[al].width = 25
                ws.column_dimensions['B'].width = 20
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 20
                ws.column_dimensions['H'].width = 20
                ws.column_dimensions['I'].width = 20
                ws.column_dimensions['J'].width = 20
                ws.column_dimensions['K'].width = 20
                ws.column_dimensions['L'].width = 20
                ws.column_dimensions['M'].width = 20
                ws.column_dimensions['N'].width = 20
                ws.column_dimensions['O'].width = 20
                ws.column_dimensions['P'].width = 20
                ws.column_dimensions['Q'].width = 20
                ws.column_dimensions['R'].width = 20
                ws.column_dimensions['S'].width = 20
                ws.column_dimensions['T'].width = 20
                ws.column_dimensions['U'].width = 20
                ws.column_dimensions['V'].width = 20
                ws.column_dimensions['W'].width = 20
                ws.column_dimensions['X'].width = 20
                ws.column_dimensions['Y'].width = 20
                ws.column_dimensions['Z'].width = 20

        # Excel Logo display
        cwd = os.path.abspath(__file__)
        path = cwd.rsplit('/', 2)
        img_path = path[0] + '/static/src/img/logo.png'
        img = openpyxl.drawing.image.Image(img_path)
        ws.add_image(img, 'A2')
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:A5')
        ws.merge_cells('B1:B5')
        ws.merge_cells('A1:B1')
        ws.merge_cells('A2:B2')
        ws.merge_cells('A3:B3')
        ws.merge_cells('A4:B4')
        ws.merge_cells('A5:B5')

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data

    @api.multi
    def generate_pdf_report(self):
        return self.env.ref('pappaya_marketing.report_address_collection').get_report_action(self)

    @api.multi
    def generate_excel_report(self):
        data = base64.encodestring(self.generate_xl())
        attach_vals = {
            'name': '%s.xls' % ('Address Collection - '+ self.name),
            'datas': data,
            'datas_fname': '%s.xls' % ('Address Collection - '+ self.name),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }


class AddressCollectionBranch(models.TransientModel):
    _name = "address.collection.branch"

    s_no = fields.Char(string='S.No',size=100)
    report_type = fields.Selection([('branch','Branch Report'),('branch_employee','Branch & Employee Report'),('employee','Employee Report')], string='Report Type')
    address_branch_id = fields.Many2one('address.collection', string='Address')
    branch_id = fields.Many2one('res.company', string='Branch Name')
    course_id = fields.Many2one('pappaya.lead.course', string='Lead Course')
    address_count = fields.Char(string='Address Count',size=100)
    course1 = fields.Char('Course 1',size=30)
    course2 = fields.Char('Course 2',size=30)
    course3 = fields.Char('Course 3',size=30)
    course4 = fields.Char('Course 4',size=30)
    course5 = fields.Char('Course 5',size=30)
    course6 = fields.Char('Course 6',size=30)
    course7 = fields.Char('Course 7',size=30)
    course8 = fields.Char('Course 8',size=30)
    course9 = fields.Char('Course 9',size=30)
    course10 = fields.Char('Course 10',size=30)
    course11 = fields.Char('Course 11',size=30)
    course12 = fields.Char('Course 12',size=30)
    course13 = fields.Char('Course 13',size=30)
    course14 = fields.Char('Course 14',size=30)
    course15 = fields.Char('Course 15',size=30)

class AddressCollectionBranchEmployee(models.TransientModel):
    _name = "address.collection.branch.employee"

    s_no = fields.Char(string='S.No',size=100)
    report_type = fields.Selection([('branch','Branch Report'),('branch_employee','Branch & Employee Report'),('employee','Employee Report')], string='Report Type')
    address_branch_employee_id = fields.Many2one('address.collection', string='Address')
    branch_id = fields.Many2one('res.company', string='Branch Name')
    employee_id = fields.Many2one('hr.employee', string='Employee Name')
    emp_id = fields.Char(string='Employee ID',size=10)
    emp_mobile = fields.Char(string='Mobile Number', size=10)
    course_id = fields.Many2one('pappaya.lead.course', string='Lead Course')
    address_count = fields.Char(string='Address Count',size=100)
    course1 = fields.Char('Course 1',size=50)
    course2 = fields.Char('Course 2',size=50)
    course3 = fields.Char('Course 3',size=50)
    course4 = fields.Char('Course 4',size=50)
    course5 = fields.Char('Course 5',size=50)
    course6 = fields.Char('Course 6',size=50)
    course7 = fields.Char('Course 7',size=50)
    course8 = fields.Char('Course 8',size=50)
    course9 = fields.Char('Course 9',size=50)
    course10 = fields.Char('Course 10',size=50)
    course11 = fields.Char('Course 11',size=50)
    course12 = fields.Char('Course 12',size=50)
    course13 = fields.Char('Course 13',size=50)
    course14 = fields.Char('Course 14',size=50)
    course15 = fields.Char('Course 15',size=50)

class AddressCollectionEmployee(models.TransientModel):
    _name = "address.collection.employee"

    s_no = fields.Char(string='S.No',size=50)
    report_type = fields.Selection([('branch','Branch Report'),('branch_employee','Branch & Employee Report'),('employee','Employee Report')], string='Report Type')
    address_employee_id = fields.Many2one('address.collection', string='Address')
    branch_id = fields.Many2one('res.company', string='Branch Name')
    employee_id = fields.Many2one('hr.employee', string='Employee Name')
    emp_id = fields.Char(string='Employee ID',size=10)
    emp_mobile = fields.Char(string='Mobile Number', size=10)
    course_id = fields.Many2one('pappaya.lead.course', string='Lead Course')
    address_count = fields.Char(string='Address Count',size=50)
    course1 = fields.Char('Course 1',size=50)
    course2 = fields.Char('Course 2',size=50)
    course3 = fields.Char('Course 3',size=50)
    course4 = fields.Char('Course 4',size=50)
    course5 = fields.Char('Course 5',size=50)
    course6 = fields.Char('Course 6',size=50)
    course7 = fields.Char('Course 7',size=50)
    course8 = fields.Char('Course 8',size=50)
    course9 = fields.Char('Course 9',size=50)
    course10 = fields.Char('Course 10',size=50)
    course11 = fields.Char('Course 11',size=50)
    course12 = fields.Char('Course 12',size=50)
    course13 = fields.Char('Course 13',size=50)
    course14 = fields.Char('Course 14',size=50)
    course15 = fields.Char('Course 15',size=50)
    
    
# """Address Collection Import Feature"""
# """Based on XLS template given by client this feature implemented."""

class student_address_data_import_line(models.TransientModel):
    _name='student.address.data.import.line'
    _description='Sudent Adress Data import Line'
    
    sequence = fields.Char('Sequence',size=50)
    student_address_data_import_id = fields.Many2one('student.address.data.import', 'Address Import ID')
    branch_id = fields.Many2one('res.company','Adm.Branch', domain=[('country_id.is_active','=',True)])
    employee_id = fields.Many2one('hr.employee','PRO')
    name = fields.Char('Student Name',size=100)
    mobile = fields.Char('Mobile', size=10)
    studying_course_id = fields.Many2one('pappaya.lead.course', string="Studying Course")
    
    city = fields.Char('City',size=50)
    mandal = fields.Char('Mandal',size=50)
    ward = fields.Char('Ward',size=50)
    area = fields.Char('Area',size=50)
    village = fields.Char('Village',size=50)
    studying_school_name = fields.Char(string='School Name',size=100)
    
    father_name = fields.Char('Father',size=100)
#     gender = fields.Selection([('male', 'Male'), ('female', 'Female')], 'Gender')
    location_type = fields.Selection([('urban','Urban'), ('rural','Rural')], string='Location Type', default='urban')
    state_id  = fields.Many2one('res.country.state', string='State')
    district_id = fields.Many2one('state.district', string='District')
    
#     city_id = fields.Many2one('pappaya.city', string='City')
#     ward_id = fields.Many2one('pappaya.ward', string='Ward')
#     area_id = fields.Many2one('pappaya.area', string='Area')
#     mandal_id = fields.Many2one('pappaya.mandal.marketing', string='Mandal')
#     village_id = fields.Many2one('pappaya.village', string='Village')
    
    pincode = fields.Char('Pin Code', size=6)    
    status = fields.Char('Status')
    
    @api.model
    def default_get(self, fields):
        res = super(student_address_data_import_line, self).default_get(fields)
        res['branch_id'] = self._context.get('branch_id') or False
        res['employee_id'] = self._context.get('employee_id') or False
        return res
    
    @api.onchange('state_id')
    def onchange_state_id(self):
        domain = {}
        domain['state_id'] = [('id','in',[])]
        state_ids = self.env['res.country.state'].search([('country_id','in',self.env['res.country'].search([('is_active','=',True)]).ids)])
        if state_ids:
            domain['state_id'] = [('id','in',state_ids.ids)]
        return {'domain':domain}   

    @api.onchange('state_id')
    def _onchange_state_id(self):
        if self.state_id:
            self.district_id = ''
    # End

    @api.onchange('location_type')
    def _onchange_location_type(self):
        if self.location_type:
            self.city = ''; self.ward = ''; self.area = '';
            self.mandal = ''; self.village = ''

            
class fileWiseImportedDetails(models.TransientModel):
    _name='file.wise.imported.details'    
    _description='File Wise Imported Details'
    
    student_address_data_import_id = fields.Many2one('student.address.data.import', 'Import ID')
    file_name = fields.Char('Imported File Name',size=50)
    valid = fields.Integer('Valid')
    invalid = fields.Integer('Invalid')
    
    
class student_address_data_import(models.TransientModel):
    _name='student.address.data.import'
    _description='Sudent Adress Data import'
    _rec_name='file_name'
    
    payroll_branch_id = fields.Many2one('pappaya.payroll.branch', 'Payroll Branch')
    employee_id = fields.Many2one('hr.employee', 'PRO Employee')
    branch_type = fields.Selection([('school','School'),('college','College')], 'School/College', default='college')
    branch_id = fields.Many2one('res.company', 'Adm.Branch')
    file = fields.Binary('Data File')
    file_name = fields.Char('File Name',size=50)
    student_address_data_import_ids = fields.One2many('student.address.data.import.line', 'student_address_data_import_id', 'Student Address')
    file_wise_imported_details_ids = fields.One2many('file.wise.imported.details', 'student_address_data_import_id', 'Files Imported')
    
    import_template_file = fields.Binary('Import Template')
    import_template_file_name = fields.Char('Import template file name',size=200)
    import_wizard_status = fields.Char('Status', default="Address imported successfully.",size=200)
    
    @api.multi 
    def download_import_template(self):
        return {
            'type' : 'ir.actions.act_url',
            'url': '/web/binary/download_document?model=student.address.data.import&field=import_template_file&id=%s&filename=excel_template.xlsx' %(self.id),
            'target': 'self',
        }

    @api.model
    def default_get(self, fields):
        res = super(student_address_data_import, self).default_get(fields)
        import_template = self.env['address.import.template'].search([], limit=1)
        if import_template:
            res['import_template_file'] = import_template.file
            res['import_template_file_name'] = import_template.file_name
        return res
    
    @api.onchange('payroll_branch_id', 'branch_id','branch_type')
    def onchange_payroll_branch_id(self):
        domain = {}; domain['branch_id']=[('id','in',[])]
        if self.payroll_branch_id:
            domain['branch_id'] = [('id','in',self.env['res.company'].search([('id','!=',1),('type','=','branch'),('branch_type','=',self.branch_type),('tem_state_id','=',self.payroll_branch_id.state_id.id)]).ids)]
        return {'domain':domain}
    
    @api.onchange('employee_id')
    def onchange_employee_id(self):
        if self.employee_id:
            mapped_files = []
            file_wise_imported_details_ids = []
            for record in self.env['pappaya.lead.stud.address'].search([('employee_id','=',self.employee_id.id),('file_name','!=',False)]):
                file_info = {}
                file_info.update({
                    'file_name':record.file_name,
                    'valid':0,
                    'invalid':0
                    })
                if record.status == 'valid':
                    file_info['valid'] += 1
                else:
                    file_info['invalid'] += 1
                    file_wise_imported_details_ids.append((0,0,file_info))
            if file_wise_imported_details_ids:
                self.file_wise_imported_details_ids = file_wise_imported_details_ids
    
    @api.multi
    def import_address_data(self):
        for line in self.student_address_data_import_ids:
            if line.branch_id and line.employee_id and line.name and line.mobile and line.studying_course_id and line.state_id and line.district_id and line.studying_school_name:                            
                if not self.env['pappaya.lead.stud.address'].search([('name','=',line.name),('mobile','=',line.mobile),('studying_course_id','=',line.studying_course_id.id)]):
                    data_dict = {
                        'branch_id': line.branch_id.id,
                        'employee_id':line.employee_id.id,
                        'name':line.name,
                        'mobile':line.mobile,
                        'studying_course_id':line.studying_course_id.id,
                        'studying_school_name':line.studying_school_name,
                        'father_name':line.father_name,
#                         'gender':line.gender,
                        'location_type':line.location_type,
                        'state_id':line.state_id.id,
                        'district_id':line.district_id.id,
                        'city':line.city,
                        'ward':line.ward,
                        'area':line.area,
                        'mandal':line.mandal,
                        'village':line.village,
                        'pincode':line.pincode,
                        'file_name':self.file_name,
                        }
                    
                    invalid_mobile_numbers = ['0000000000','1111111111','2222222222','3333333333','4444444444','5555555555','6666666666',
                                              '7777777777','8888888888','9999999999']
                    match_mobile = re.match('^[\d]*$', line.mobile)
                    if not match_mobile or len(line.mobile) != 10 or line.mobile in invalid_mobile_numbers:
                        line.status = 'Invalid Mobile' 
                    if line.pincode:
                        invalid_pincode_list = ['000000','111111','222222','333333','444444','555555','666666','777777',
                                        '888888','999999']
                        match_zip_code = re.match('^[\d]*$', line.pincode)
                        if line.pincode in invalid_pincode_list or not match_zip_code or len(line.pincode) != 6:
                            line.status = 'Invalid Mobile'      
                    try:
                        new_id = self.env['pappaya.lead.stud.address'].create(data_dict)
                        line.sequence = new_id.sequence
                        line.status = new_id.status
                    except:
                        self._cr.rollback()
                        self.status = 'Please provide valid input.'
                else:
                    line.status = 'Duplicate entry.'
            else:
                line.status = 'Fill all mandatory fields.'
        return True
    
    @api.multi
    def process_file(self):
        if not self.file:
            raise ValidationError("Please upload file and try again.")
        if not self.payroll_branch_id or not self.employee_id or not self.branch_id:
            raise ValidationError("Payroll branch, PRO and Adm. branch should be mandatory to process.")
        
        if self.file:
            self.student_address_data_import_ids = False
            file = base64.b64decode(bytes(self.file))
            excel_fileobj = TemporaryFile('wb+')
            excel_fileobj.write(file)
            excel_fileobj.seek(0)
            workbook = openpyxl.load_workbook(excel_fileobj, data_only=True)
            line_ids = []
            address_need_to_be_corrected= []
            for sheet in workbook.worksheets:
                row_count = 0
                for row in sheet.rows:
                    row_count += 1
                    if row_count == 1:
                        record_correction_header_row = {}
                        for column in range(sheet.max_column):
                            record_correction_header_row.update({ column : str(row[column].value).strip()})
                        address_need_to_be_corrected.append(record_correction_header_row)
                    if row_count > 1:
                        data_dict = {}
                        student_name='';mobile='';studying_course_id=False
                        state_id = False;district_id = False;
#                         city_id=False;ward_id=False;area_id=False;mandal_id=False
#                         school_state_id = False; school_district_id=False;school_mandal_id=False;lead_school_id=False
                        location_type=''
                        branch_id=self.branch_id.id;employee_id=self.employee_id.id
                        data_dict.update({'branch_id':branch_id, 'employee_id':employee_id})
#                         # Branch
#                         if row[1].value != "":
#                             branch_name = (str(row[1].value).strip()).upper()
#                             admission_branch = self.env['res.company'].sudo().search([('name','=',branch_name)])
#                             if admission_branch:
#                                 branch_id = admission_branch.id        
#                                 data_dict.update({'branch_id':branch_id})
#                         # PRO
#                         if row[2].value != "":
#                             emp_id = str(row[2].value).strip()
#                             employee_obj = self.env['hr.employee'].sudo().search([('emp_id','=',emp_id)])
#                             if employee_obj:
#                                 employee_id = employee_obj.id
#                                 data_dict.update({'employee_id':employee_id})                       
                        
                        # Student Name
                        if row[1].value != "" and row[1].value != None:
                            student_name = (str(row[1].value).strip()).capitalize()
                            data_dict.update({'name':student_name})    
                        
                        # Gender
#                         if row[2].value != "":
#                             gender = str(row[2].value).strip().lower()
#                             data_dict.update({'gender':gender})                        
#                         
                        # Father Name
                        if row[2].value != "" and row[2].value != None:
                            father_name = (str(row[2].value).strip()).capitalize()
                            data_dict.update({'father_name':father_name})                        
                        
                        # Location Type
                        if row[3].value != "" and row[3].value != None:
                            location_type = str(row[3].value).strip().lower()
                            data_dict.update({'location_type':location_type})                      
                        
                        # state
                        if row[4].value != "" and row[4].value != None:
                            state_name = str(row[4].value).strip()
                            state_obj = self.env['res.country.state'].sudo().search([('name','=',state_name)])
                            if state_obj:
                                state_id = state_obj.id
                                data_dict.update({'state_id':state_id})
                        
                        # District
                        if row[5].value != "" and row[5].value != None:
                            district_name = str(row[5].value).strip()
                            if state_id:
                                district_obj = self.env['state.district'].sudo().search([('name','=',district_name),('state_id','=',state_id)])
                                if district_obj:
                                    district_id = district_obj.id
                                    data_dict.update({'district_id':district_id})                
                        
                        # CITY
                        if row[6].value != "" and row[6].value != None:
                            city_name = str(row[6].value).strip()
                            data_dict.update({'city':city_name})
                            
#                             if district_id:
#                                 city_obj = self.env['pappaya.city'].sudo().search([('name','=',city_name),('district_id','=',district_id)])
#                                 if city_obj:
#                                     city_id = city_obj.id
#                                     data_dict.update({'city_id':city_id})
                        
                        # WARD
                        if row[7].value != "" and row[7].value != None:
                            if location_type == 'rural':
                                mandal_name = str(row[7].value).strip()
                                data_dict.update({'mandal':mandal_name})
                            else:
                                ward_name = str(row[7].value).strip()
                                data_dict.update({'ward':ward_name})                                
                            
#                             if city_id:
#                                 ward_obj = self.env['pappaya.ward'].sudo().search([('name','=',ward_name),('city_id','=',city_id)])
#                                 if ward_obj:
#                                     ward_id = ward_obj.id
#                                     data_dict.update({'ward_id':ward_id})   
                        
                        # AREA
                        if row[8].value != "" and row[8].value != None:
                            if location_type == 'rural':
                                village_name = str(row[8].value).strip()
                                data_dict.update({'village':village_name})
                            else:
                                area_name = str(row[8].value).strip()
                                data_dict.update({'area':area_name})                              
                        
#                             area_name = str(row[8].value).strip()
#                             data_dict.update({'area':area_name})
#                             if ward_id:
#                                 area_obj = self.env['pappaya.area'].sudo().search([('name','=',area_name),('ward_id','=',ward_id)])
#                                 if area_obj:
#                                     area_id = area_obj.id
#                                     data_dict.update({'area_id':area_id})                                
                        
#                         # MANDAL
#                         if row[9].value != "":
#                             mandal_name = str(row[9].value).strip()
#                             data_dict.update({'mandal':mandal_name})
#                             
# #                             if district_id:
# #                                 mandal_obj = self.env['pappaya.mandal.marketing'].sudo().search([('name','=',mandal_name),('district_id','=',district_id)])
# #                                 if mandal_obj:
# #                                     mandal_id = mandal_obj.id     
# #                                     data_dict.update({'mandal_id':mandal_id})
#                         
#                         # VILLAGE
#                         if row[10].value != "":
#                             village_name = str(row[10].value).strip()
#                             
#                             if mandal_id:
#                                 village_obj = self.env['pappaya.village'].sudo().search([('name','=',village_name),('mandal_id','=',mandal_id)])
#                                 if village_obj:
#                                     village_id = village_obj.id
#                                     data_dict.update({'village_id':village_id})                                              
                        
                        # MOBILE
                        if row[9].value != "" and row[9].value != None:
                            mobile = str(row[9].value).strip()
                            data_dict.update({'mobile':mobile})
                        
                        # PINCODE
                        if row[10].value != "" and row[10].value != None:
                            pincode = str(row[10].value).strip()
                            data_dict.update({'pincode':pincode})
                            
                        # SCHOOL NAME
                        if row[11].value != "" and row[11].value != None:
                            school_name = str(row[11].value).strip()
                            data_dict.update({'studying_school_name':school_name})                            
                        
#                         # SCHOOL STATE
#                         if row[14].value != "":
#                             school_state = str(row[14].value).strip()
#                             school_state_obj = self.env['res.country.state'].sudo().search([('name','=',school_state)])
#                             if school_state_obj:
#                                 school_state_id = school_state_obj.id
#                                 data_dict.update({'school_state_id':school_state_id})
#                         
#                         # SCHOOL DISTRICT
#                         if row[15].value != "":
#                             school_district = str(row[15].value).strip()
#                             if school_state_id:
#                                 school_district_obj = self.env['state.district'].sudo().search([('name','=',school_district),('state_id','=',school_state_id)])
#                                 if school_district_obj:
#                                     school_district_id = school_district_obj.id
#                                     data_dict.update({'school_district_id':school_district_id})                        
#                         
#                         # SCHOOL MANDAL
#                         if row[16].value != "":
#                             school_mandal = str(row[16].value).strip()
#                             if school_district_id:
#                                 school_mandal_obj = self.env['pappaya.mandal.marketing'].sudo().search([('name','=',school_mandal),('district_id','=',school_district_id)])
#                                 if school_mandal_obj:
#                                     school_mandal_id = school_mandal_obj.id
#                                     data_dict.update({'school_mandal_id':school_mandal_id})                        
#                         
#                         # LEAD SCHOOL
#                         if row[17].value != "":
#                             lead_school = str(row[17].value).strip()
#                             if school_state_id and school_district_id and school_mandal_id:
#                                 lead_school_obj = self.env['pappaya.lead.school'].sudo().search([('name','=',lead_school),
#                                                                                             ('state_id','=',school_state_id),
#                                                                                             ('district_id','=',school_district_id),
#                                                                                             ('mandal_id','=',school_mandal_id)
#                                                                                             ])
#                                 if lead_school_obj:
#                                     lead_school_id = lead_school_obj.id
#                                     data_dict.update({'lead_school_id':lead_school_id})
#                                 
#                         # LEAD SCHOOL
                        if row[12].value != "" and row[12].value != None:
                            studying_course = (str(row[12].value).strip()).upper()
                            lead_course_obj = self.env['pappaya.lead.course'].sudo().search([('name','=',studying_course)])
                            if lead_course_obj:
                                studying_course_id = lead_course_obj.id
                                data_dict.update({'studying_course_id':studying_course_id})
                        
                        line_ids.append((0,0, data_dict))
                        
#                         if branch_id and employee_id and student_name and mobile and studying_course_id:
#                             if not self.env['pappaya.lead.stud.address'].sudo().search([('name','=',student_name),('mobile','=',mobile),('studying_course_id','=',studying_course_id)]):
#                                 line_ids.append((0,0, data_dict))
                                
            if line_ids:
                self.student_address_data_import_ids = line_ids
                
#                 ack_id = self.env['ir.model.data'].get_object_reference('pappaya_marketing', 'view_student_address_data_import_form')[1] or False
#                 return {
#                     'name': _('Address Import Wizard'),
#                     'type': 'ir.actions.act_window',
#                     'view_type': 'form',
#                     'view_mode': 'form',
#                     'res_id': self.id,
#                     'res_model': 'student.address.data.import',
#                     'views': [(ack_id, 'form')],
#                     'view_id': ack_id,
#                     'target': 'current',
#                     'context':self._context
#                 }
#                                 try:
#                                     self.env['pappaya.lead.stud.address'].create(data_dict)
#                                 except:
#                                     self._cr.rollback()
#                                     address_to_be_corrected_dict = {}
#                                     for column in range(sheet.max_column):
#                                         address_to_be_corrected_dict.update({ column : str(row[column].value).strip()})
#                                     address_need_to_be_corrected.append(address_to_be_corrected_dict)
#                             else:
#                                 address_to_be_corrected_dict = {}
#                                 for column in range(sheet.max_column):
#                                     address_to_be_corrected_dict.update({ column : str(row[column].value).strip()})
#                                 address_need_to_be_corrected.append(address_to_be_corrected_dict)
#                         else:
#                             address_to_be_corrected_dict = {}
#                             for column in range(sheet.max_column):
#                                 address_to_be_corrected_dict.update({ column : str(row[column].value).strip()})
#                             address_need_to_be_corrected.append(address_to_be_corrected_dict)
#                 if address_need_to_be_corrected and len(address_need_to_be_corrected) > 1:
#                     f_name= 'address_to_be_corrected.xlsx'
#                     workbook = Workbook()
#                     worksheet1 = workbook.active
#                     worksheet1.title = "Address Collection"
#                     ft1 = Font(size=11, bold=True)
#                     header1 = NamedStyle(name="header1", font=ft1)
#                     worksheet1.row_dimensions[1].height = 24
#                     for record_to_be_correct in address_need_to_be_corrected:
#                         list = []
#                         for key in range(len(record_to_be_correct)):
#                             list.append(record_to_be_correct[key])
#                         if list:
#                             worksheet1.append(list)    
#                     for cell in worksheet1["1:1"]:
#                         cell.style = header1
#                     out = BytesIO()
#                     workbook.save(out)
#                     out.seek(0)
#                     data = out.read()
#                     data = base64.encodestring(data)
#                     """returning the output xls as binary"""
#                     self.write({'records_to_be_correct_file':data, 'records_to_be_correct_file_name':f_name})
#                     out.close()
#                     workbook.close()
#             
#             ack_id = self.env['ir.model.data'].get_object_reference('pappaya_marketing', 'view_address_import_acknowledge_form')[1] or False
#             return {
#                 'name': _('Address Import Wizard'),
#                 'type': 'ir.actions.act_window',
#                 'view_type': 'form',
#                 'view_mode': 'form',
#                 'res_id': self.id,
#                 'res_model': 'student.address.data.import',
#                 'views': [(ack_id, 'form')],
#                 'view_id': ack_id,
#                 'target': 'new',
#                 'context':self._context
#             }               
            
                
class UserInfo(models.Model):
    _name = 'user.info'

    name = fields.Char("Name",size=100)
    mobile = fields.Char("Mobile No",size=10)

    _rec_name = 'name'

    @api.model
    def create(self, vals):
        print ("\n\nCreate method called", vals, "\n\n")
        return super(UserInfo, self).create(vals)

UserInfo()
            