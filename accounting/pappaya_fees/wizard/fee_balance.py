from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from datetime import datetime, date
from odoo.tools.misc import DEFAULT_SERVER_DATE_FORMAT
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import base64
import os
from io import BytesIO
import logging
_logger = logging.getLogger(__name__)

class FeeBalanceWizard(models.TransientModel):
    _name = "fee.balance.wizard"

    society_ids =fields.Many2many('operating.unit','society_balance_strength_rel','company_id','society_id',string= 'Society')
    school_ids = fields.Many2many('operating.unit','school_balance_rel','company_id','school_id', string='Branch')
    academic_year_ids = fields.Many2many('academic.year','academic_year_balance_rel','school_id', 'academic_year_id', string='Academic Year')
    grade_ids = fields.Many2many('pappaya.grade','grade_balance_rel','school_id', 'grade_id', string='Grade')

    @api.onchange('society_ids')
    def onchange_society_ids(self):
        if self.society_ids:
            self.school_ids = []
            return {'domain': {'school_ids': [('type', '=', 'school'),('parent_id', 'in', self.society_ids.ids)]}}
    
    @api.onchange('school_ids')
    def onchange_school_id(self):
        if self.school_ids:
            year_ids = self.academic_ids()
            return {'domain': {'academic_year_ids': [('id', 'in', year_ids)]}}

    @api.multi    
    def academic_ids(self):
        date_list = []
        year_ids = []
        if self.school_ids:
            max_start_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="start_date desc",limit=1).start_date
            min_start_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="start_date",limit=1).start_date
            max_end_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="end_date desc",limit=1).end_date
            min_end_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="end_date",limit=1).end_date
            date_list.append(max_start_date)
            date_list.append(min_start_date)
            date_list.append(max_end_date)
            date_list.append(min_end_date)
            min_year = (datetime.strptime(min(date_list), DEFAULT_SERVER_DATE_FORMAT)).year
            max_year = (datetime.strptime(max(date_list), DEFAULT_SERVER_DATE_FORMAT)).year 
            academic_year_ids = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)])
            for academic_year in  academic_year_ids:
                if (datetime.strptime(academic_year.start_date, DEFAULT_SERVER_DATE_FORMAT)).year >= min_year or (datetime.strptime(academic_year.start_date, DEFAULT_SERVER_DATE_FORMAT)).year <= max_year:
                    year_ids.append(academic_year.id)
            return year_ids

    @api.multi
    def get_school_header(self):
        school_list = []
        if (len(self.school_ids) == 1 and len(self.society_ids) == 1) or (len(self.school_ids) == 1 and not self.society_ids):
            vals = {}
            vals['school_id'] = self.school_ids.name
            vals['logo'] = self.school_ids.logo if self.school_ids.logo else ''
            vals['street'] = self.school_ids.street if self.school_ids.street else ''
            vals['street2'] = self.school_ids.street2 if self.school_ids.street2 else ''
            vals['city'] = self.school_ids.city if self.school_ids.city else ''
            vals['zip'] = self.school_ids.zip if self.school_ids.zip else ''
            vals['phone'] = self.school_ids.phone if self.school_ids.phone else ''
            vals['fax'] = self.school_ids.fax_id if self.school_ids.fax_id else ''
            vals['email'] = self.school_ids.email if self.school_ids.email else ''
            vals['website'] = self.school_ids.website if self.school_ids.website else ''
            school_list.append(vals)
        if (len(self.society_ids) == 1 and not self.school_ids):
            vals = {}
            vals['school_id'] = self.society_ids.name
            vals['logo'] = self.society_ids.logo
            vals['street'] = self.society_ids.street if self.society_ids.street else ''
            vals['street2'] = self.society_ids.street2 if self.society_ids.street2 else ''
            vals['city'] = self.society_ids.city if self.society_ids.city else ''
            vals['zip'] = self.society_ids.zip if self.society_ids.zip else ''
            vals['phone'] = self.society_ids.phone if self.society_ids.phone else ''
            vals['fax'] = self.society_ids.fax_id if self.society_ids.fax_id else ''
            vals['email'] = self.society_ids.email if self.society_ids.email else ''
            vals['website'] = self.society_ids.website if self.society_ids.website else ''
            school_list.append(vals)
        if (len(self.society_ids) == 1 and len(self.school_ids) != 1):
            vals = {}
            vals['school_id'] = self.society_ids.name
            vals['logo'] = self.society_ids.logo
            vals['street'] = self.society_ids.street if self.society_ids.street else ''
            vals['street2'] = self.society_ids.street2 if self.society_ids.street2 else ''
            vals['city'] = self.society_ids.city if self.society_ids.city else ''
            vals['zip'] = self.society_ids.zip if self.society_ids.zip else ''
            vals['phone'] = self.society_ids.phone if self.society_ids.phone else ''
            vals['fax'] = self.society_ids.fax_id if self.society_ids.fax_id else ''
            vals['email'] = self.society_ids.email if self.society_ids.email else ''
            vals['website'] = self.society_ids.website if self.society_ids.website else ''
            school_list.append(vals)
        return school_list

    @api.multi
    def get_society_header(self):
        society_list = []
        if not self.society_ids and not self.school_ids:
            soc_list = ''
            obj = self.env['res.company'].search([('type', '=', 'society')])
            for record in obj:
                soc_list += str(record.name) + ', '
            vals = {}
            vals['society_id'] = soc_list[:-2]
            society_list.append(vals)
        else:
            sc_list = ''
            for record in self.society_ids:
                sc_list += str(record.name) + ', '
            vals = {}
            vals['society_id'] = sc_list[:-2]
            society_list.append(vals)
        return society_list

    @api.multi
    def get_head(self):
        head_list, head_dict = [],{}
        for obj in self.env['pappaya.fees.term.line'].search([],order="name asc"):
            if obj.name not in head_dict:
                head_dict.update({obj.name: [obj]})
            else:
                head_dict[obj.name].append(obj)
        for key in sorted(head_dict):
            vals = {}
            vals['header'] = key
            head_list.append(vals)
        return head_list

    @api.multi
    def get_balance_data(self):

        domain, data_list, conc_dict = [], [], {}
        if self.society_ids:
            domain.append(('society_id','in',self.society_ids.ids))
        if self.school_ids:
            domain.append(('school_id','in',self.school_ids.ids))
        if self.academic_year_ids:
            domain.append(('academic_year_id','in',self.academic_year_ids.ids))
        if self.grade_ids:
            domain.append(('grade_id','in',self.grade_ids.ids))
        fee_obj = self.env['pappaya.fees.collection'].search([('fees_collection_line.term_state','=','due')]+domain)
        for obj in fee_obj:
            if obj.bulk_term_state == 'due':
                vals = {}
                head_dict = {}
                for head in self.env['pappaya.fees.term.line'].search([],order='name asc'):
                    if head.name not in head_dict:
                        head_dict.update({head.name: [head]})
                    else:
                        head_dict[head.name].append(head)
                amt_list, total_amt = [], 0.0
                vals['fee'] = []
                for i in sorted(head_dict):
                    srch_res = []
                    stud_fee = self.env['student.fees.collection'].search([('collection_id','=',obj.id),('term_state','=','due'),('name', '=', i)])
                    srch_res.append(stud_fee)
                    for fee_struc in srch_res:
                        total_amt += fee_struc.due_amount
                        fee_vals = {}
                        if fee_struc and fee_struc.due_amount == 0.0:
                            amt_list.append('%.2f' % fee_struc.due_amount or 0.0)
                            fee_vals['fee']= ('%.2f' % fee_struc.due_amount or 0.0)
                        elif fee_struc and fee_struc.due_amount > 0.0:
                            amt_list.append('%.2f' % fee_struc.due_amount or 0.0)
                            fee_vals['fee'] = ('%.2f' % fee_struc.due_amount or 0.0)
                        else:
                            amt_list.append('%.2f' % 0.0)
                            fee_vals['fee']= ('%.2f' % 0.0)
                        vals['fee'].append(fee_vals)
                vals['society'] = obj.society_id.name if obj.society_id else ''
                vals['school'] = obj.school_id.name if obj.school_id else ''
                vals['academic'] = obj.academic_year_id.name if obj.academic_year_id else ' '
                vals['grade'] = obj.grade_id.name if obj.grade_id else ''
                vals['stud_id'] = obj.student_id.enrollment_num if obj.student_id.enrollment_num else ''
                vals['stud_name'] = str(obj.student_id.name or '') + ' ' + str(obj.student_id.middle_name or '') + ' ' + str(obj.student_id.last_name or '')
                vals['parent'] = obj.student_id.father_name if obj.student_id.father_name else '' or obj.student_id.mother_name if obj.student_id.mother_name else ''
                vals['mobile'] = obj.student_id.father_mobile_no if obj.student_id.father_mobile_no else ''
                vals['date'] = 0
                vals['fee_head'] = amt_list
                vals['total'] = '%.2f' % total_amt
                data_list.append(vals)
        return data_list

    @api.multi
    def generate_xl(self):
        wb= Workbook()
        ws= wb.active
        ws.title = "FEE BALANCE LIST"
        if (len(self.school_ids) == 1 and len(self.society_ids) == 1) or (len(self.school_ids) == 1 and not self.society_ids):
            ws.append([(self.school_ids.name if self.school_ids.name else '')])
            ws.append([(self.school_ids.street if self.school_ids.street else '') + ', ' + (self.school_ids.street2 if self.school_ids.street2 else '') + ', ' + (self.school_ids.city if self.school_ids.city else '')])
            ws.append(['Tel: ' + (self.school_ids.phone if self.school_ids.phone else '') + ', ' + 'Fax: ' + (self.school_ids.fax_id if self.school_ids.fax_id else '') + ', ' + 'Email: ' + (self.school_ids.email if self.school_ids.email else '')])
            ws.append([self.school_ids.website if self.school_ids.website else ''])
            ws.append([])
        if (len(self.society_ids) == 1 and not self.school_ids) or (len(self.society_ids) == 1 and len(self.school_ids) > 1):
            ws.append([(self.society_ids.name if self.society_ids.name else '')])
            ws.append([(self.society_ids.street if self.society_ids.street else '') + ', ' + (self.society_ids.street2 if self.society_ids.street2 else '') + ', ' + (self.society_ids.city if self.society_ids.city else '')])
            ws.append(['Tel: ' + (self.society_ids.phone if self.society_ids.phone else '') + ', ' + 'Fax: ' + (self.society_ids.fax_id if self.society_ids.fax_id else '') + ', ' + 'Email: ' + (self.society_ids.email if self.society_ids.email else '')])
            ws.append([self.society_ids.website if self.society_ids.website else ''])
            ws.append([])
        if (not self.society_ids and not self.school_ids) or (len(self.school_ids) > 1 and not self.society_ids):
            soc_list = ''
            obj = self.env['res.company'].search([('type', '=', 'society')])
            for record in obj:
                soc_list += str(record.name) + ', '
            ws.append([])
            ws.append([])
            ws.append([soc_list[:-2]])
            ws.append([])
            ws.append([])
        if (len(self.society_ids) > 1) or (len(self.society_ids) > 1 and len(self.school_ids) > 1):
            sc_list = ''
            for record in self.society_ids:
                sc_list += str(record.name) + ', '
            ws.append([])
            ws.append([])
            ws.append([sc_list[:-2]])
            ws.append([])
            ws.append([])
        ws.append([])
        ws.append(['FEE BALANCE LIST'])
        ws.append(['As on ' + (datetime.now()).strftime('%d/%b/%Y')])
        ws.append([])
        head_list, head_dict = [], {}
        for obj in self.env['pappaya.fees.term.line'].search([], order="name asc"):
            if obj.name not in head_dict:
                head_dict.update({obj.name: [obj]})
            else:
                head_dict[obj.name].append(obj)
        for key in sorted(head_dict):
            head_list.append(key)
        head_count = len(head_list)
        ws.append(['Society','Branch','Academic Year','Class','Student ID','Student Name','Mobile No']+head_list+['Total'])
        t_count = 10
        # Fetch data
        for xl_list in self.get_balance_data():
            ws.append([xl_list['society'],xl_list['school'],xl_list['academic'],xl_list['grade'],xl_list['stud_id'],xl_list['stud_name'],
                       xl_list['mobile']] + xl_list['fee_head'] + [xl_list['total']])
            alphabets = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                         'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
            center = 0
            for al in alphabets:
                center += 1
                ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
            t_count += 1
        #Company Details
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=15, bold=True)
        header1 = NamedStyle(name="header1", font=ft1)
        ws['A1'].style = header1
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:O1')
        ft2 = Font(size=11)
        header2 = NamedStyle(name="header2", font=ft2)
        ws['A2'].style = header2
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:O2')
        ws['A3'].style = header2
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:O3')
        ws['A4'].style = header2
        ws['A4'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A4:O4')
        ws['A5'].style = header2
        ws['A5'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A5:O5')
        ws['A6'].style = header2
        ws['A6'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A6:O6')
        # styles for row7
        ft4 = Font(size=12, bold=True)
        header3 = NamedStyle(name="header3", font=ft4)
        ws['A7'].style = header3
        ws['A7'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A7:O7')
        ws['A8'].style = header3
        ws['A8'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A8:O8')
        ft6 = Font(size=12, bold=True)
        header6 = NamedStyle(name="header6", font=ft6)
        ws['A9'].style = header6
        ws['A9'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A9:O9')
        alphabets = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T','U', 'V', 'W', 'X', 'Y', 'Z']
        al_len = 10 + head_count
        for al in alphabets:
            al_len -= 1
            ws[al + str(10)].style = header6
            ws[al + str(10)].alignment = Alignment(horizontal="center")
            ws[al + str(t_count + 1)].alignment = Alignment(horizontal="center")
        thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
        ws.cell(row=10, column=1).border = thin_border
        ws.cell(row=10, column=2).border = thin_border
        ws.cell(row=10, column=3).border = thin_border
        ws.cell(row=10, column=4).border = thin_border
        ws.cell(row=10, column=5).border = thin_border
        ws.cell(row=10, column=6).border = thin_border
        ws.cell(row=10, column=7).border = thin_border
        ws.cell(row=10, column=8).border = thin_border
        ws.cell(row=10, column=9).border = thin_border
        ws.cell(row=10, column=10).border = thin_border
        ws.cell(row=10, column=11).border = thin_border
        ws.cell(row=10, column=12).border = thin_border
        ws.cell(row=10, column=13).border = thin_border
        # Width style
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 22
        ws.column_dimensions['I'].width = 22
        ws.column_dimensions['J'].width = 22
        ws.column_dimensions['K'].width = 22
        ws.column_dimensions['L'].width = 22
        ws.column_dimensions['M'].width = 22
        ws.column_dimensions['N'].width = 22

        # cwd = os.path.abspath(__file__)
        # path = cwd.rsplit('/', 2)
        # img_path = path[0] + '/static/src/img/dis_logo.png'
        # image_path = img_path.replace("pappaya_fees", "pappaya_core")
        # img = openpyxl.drawing.image.Image(image_path)
        # ws.add_image(img, 'A1')
        # ws['A1'].alignment = Alignment(horizontal="center")
        # ws.merge_cells('A1:A5')

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data

    @api.multi
    def generate_pdf_report(self):
        if not self.get_balance_data():
            raise ValidationError(_("No record found..!"))
        if self.get_balance_data():
            return self.env['report'].get_action(self, 'pappaya_fees.template_fee_balance')

    @api.multi
    def generate_excel_report(self):
        if not self.get_balance_data():
            raise ValidationError(_("No record found..!"))
        data = base64.encodestring(self.generate_xl())
        attach_vals = {
            'name': '%s.xls' % ('Fee Balance List'),
            'datas': data,
            'datas_fname': '%s.xls' % ('Fee Balance List'),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }

FeeBalanceWizard()


