from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from datetime import datetime, date
from odoo.tools.misc import DEFAULT_SERVER_DATE_FORMAT
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import base64
import os
from io import BytesIO
import logging
_logger = logging.getLogger(__name__)

class SchoolwiseConsolidatedWizard(models.TransientModel):
    _name = "schoolwise.consolidated.wizard"

    # society_ids = fields.Many2many('res.company', 'society_consolidated_rel', 'company_id', 'society_id',string='Society')
    school_ids = fields.Many2many('operating.unit','school_consolidated_rel','company_id','school_id', string='Branch')
    academic_year_ids = fields.Many2many('academic.year','academic_year_consolidated_rel','school_id', 'academic_year_id', string='Academic Year')
    #grade_ids = fields.Many2many('pappaya.grade','grade_wise_balance_id','school_id', 'grade_id', string='Grade')

    # @api.onchange('society_ids')
    # def onchange_society_ids(self):
    #     if self.society_ids:
    #         self.school_ids = []
    #         return {'domain': {'school_ids': [('type', '=', 'school'), ('parent_id', 'in', self.society_ids.ids)]}}

    @api.onchange('school_ids')
    def onchange_school_id(self):
        if self.school_ids:
            year_ids = self.academic_ids()
            return {'domain': {'academic_year_ids': [('id', 'in', year_ids)]}}

    @api.multi    
    def academic_ids(self):
        date_list = []
        year_ids = []
        if self.school_ids:
            max_start_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="start_date desc",limit=1).start_date
            min_start_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="start_date",limit=1).start_date
            max_end_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="end_date desc",limit=1).end_date
            min_end_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="end_date",limit=1).end_date
            date_list.append(max_start_date)
            date_list.append(min_start_date)
            date_list.append(max_end_date)
            date_list.append(min_end_date)
            min_year = (datetime.strptime(min(date_list), DEFAULT_SERVER_DATE_FORMAT)).year
            max_year = (datetime.strptime(max(date_list), DEFAULT_SERVER_DATE_FORMAT)).year 
            academic_year_ids = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)])
            for academic_year in  academic_year_ids:
                if (datetime.strptime(academic_year.start_date, DEFAULT_SERVER_DATE_FORMAT)).year >= min_year or (datetime.strptime(academic_year.start_date, DEFAULT_SERVER_DATE_FORMAT)).year <= max_year:
                    year_ids.append(academic_year.id)
            return year_ids

    @api.multi
    def generate_xl(self):
        wb= Workbook()
        ws= wb.active
        ws.title = "Schoolwise Consolidated Report"
        if (len(self.school_ids) == 1):
            ws.append([(self.school_ids.name if self.school_ids.name else '')])
            ws.append([(self.school_ids.street if self.school_ids.street else '') + ', ' + (self.school_ids.street2 if self.school_ids.street2 else '') + ', ' + (self.school_ids.city if self.school_ids.city else '')])
            ws.append(['Tel: ' + (self.school_ids.mobile if self.school_ids.mobile else '') + ', ' + 'Fax: ' + (self.school_ids.fax_id if self.school_ids.fax_id else '') + ', ' + 'Email: ' + (self.school_ids.email if self.school_ids.email else '')])
            ws.append([self.school_ids.website if self.school_ids.website else ''])
            ws.append([])
        if not self.school_ids:
            soc_list = ''
            obj = self.env['res.company'].sudo().search([('type', '=', 'school')])
            for record in obj:
                soc_list += str(record.name) + ', '
            ws.append([])
            ws.append([])
            ws.append([soc_list[:-2]])
            ws.append([])
            ws.append([])
        if (len(self.school_ids) > 1):
            sc_list = ''
            for record in self.school_ids:
                sc_list += str(record.name) + ', '
            ws.append([])
            ws.append([])
            ws.append([sc_list[:-2]])
            ws.append([])
            ws.append([])
        ws.append([])
        ws.append(['SCHOOLWISE CONSOLIDATED REPORT'])
        ws.append([])
        disc_list, disc_dict = [], {}
        clt_list, clt_dict = [], {}
        domain1 = []
        if self.school_ids:
            domain1.append(('school_ids', 'in', self.school_ids.ids))
        if self.academic_year_ids:
            domain1.append(('academic_year_id', 'in', self.academic_year_ids.ids))
        for fee_term in self.env['pappaya.fees.term'].sudo().search(domain1):
            for obj in fee_term.term_divide_ids:
                if obj.name not in disc_dict:
                    disc_dict.update({(' T' + obj.name + ' Tuition Discount'): [obj]})
                else:
                    disc_dict[('T' + obj.name + ' Tuition Discount')].append(obj)

            for clt in fee_term.term_divide_ids:
                if clt.name not in clt_dict:
                    clt_dict.update({('T' + clt.name + ' Tuition Collect'): [clt]})
                else:
                    clt_dict[('T' + clt.name + ' Tuition Collect')].append(clt)
        cnt = 0
        for key in sorted(set(disc_dict)):
            cnt += 1
            disc_list.append('T' + str(cnt) + ' Tuition Charge')
            disc_list.append('T' + str(cnt) + ' Tuition Discount Type')
            disc_list.append(key)

        for key in sorted(set(clt_dict)):
            clt_list.append(key)

        ws.append(['Status','Inactive Type','Opening Balance','Debit Open','Credit Open','Student Number','RTE Status','Branch','Branch Desc',
                   'Student Name','Current Grade','Registration Fee','Registration Fee Paid','Admission Fee','Admission Fee Paid','Reg/Adm Discount Type','Reg/Adm Discount',
                   ]+disc_list+clt_list+['Refund Amount','Collection on OP Balance','Other Balance','Balance','Current Year Collection','Next Year Collection'])
        t_count = 9

        domain = []
        # if self.society_ids:
        #     domain.append(('society_id','in',self.society_ids.ids))
        if self.school_ids:
            domain.append(('school_id','in',self.school_ids.ids))
        if self.academic_year_ids:
            domain.append(('academic_year_id','in',self.academic_year_ids.ids))
        if self.grade_ids:
            domain.append(('grade_id','in',self.grade_ids.ids))
        for fees in self.env['pappaya.fees.collection'].sudo().search(['|',('active','=',True),('active','=',False)]+domain,order='id asc'):
            if fees.active == True:
                status = 'Active'
            else:
                status = 'Inactive'
            student_number = fees.enrollment_number if fees.enrollment_number else ''
            rte_status = fees.enquiry_mode if fees.enquiry_mode else ''
            school = fees.school_id.name if fees.school_id.name else ''
            school_desc = fees.school_id.parent_id.name if fees.school_id.parent_id else ''
            student_name = ((str(fees.student_id.name) + ' ' + str(fees.student_id.middle_name or '') + str(fees.student_id.last_name)) if fees.student_id else fees.applicant_name)
            grade = fees.grade_id.name
            inactive_type = ''
            state_dict = {'draft':'Enquiry','reg_process':'Registered','reject':'Rejected', 'pending':'Pending','tc':'TC',
                            'admitted':'Admitted','cancel':'Cancelled','admission_cancel':'Admission Cancel','strikeoff':'StrikeOff','transferred': 'Transferred','done': 'Transferred'}
            if fees.active == False:
                inactive_type += state_dict[fees.enquiry_id.state] if fees.enquiry_id.state else ''
            reg_fee, reg_fee_paid, adm_fee, adm_fee_paid, reg_adm_disc, refund = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
            charge_lst, disc_type_lst, disc_lst, collect_lst = [], [], [], []
            bal, curr_coll = 0.0, 0.0
            reg_fee_disc,adm_fee_disc = 0.0,0.0
            reg_adm_disc_type = ''
            t_lst = []
            loo = 0
            aa = len(clt_list)
            term_line_empty = 0
            for line in fees.fees_collection_line:
                if 'Reg' in line.name:
                    reg_fee = line.amount
                    reg_fee_paid = line.total_paid
                    reg_fee_disc += line.concession_amount
                    curr_coll += line.total_paid
                    bal += line.amount - line.concession_amount
                if 'Reg' in line.name and line.concession_type_id:
                    reg_adm_disc_type += line.concession_type_id.code
                if 'Adm' in line.name:
                    adm_fee = line.amount
                    adm_fee_paid = line.total_paid
                    adm_fee_disc += line.concession_amount
                    curr_coll += line.total_paid
                    bal += line.amount - line.concession_amount
                if 'Adm' in line.name and line.concession_type_id:
                    reg_adm_disc_type += line.concession_type_id.code
                reg_adm_disc = reg_fee_disc + adm_fee_disc
                loo += 1
                #if loo <= aa:
                if 'Term' in line.name:
                    attr_val = {}
                    attr_val['charge'] = '%.2f' % line.amount or 0.0
                    attr_val['disc_type'] = line.concession_type_id.code if line.concession_type_id.code else ''
                    attr_val['disc'] = '%.2f' % line.concession_amount or 0.0
                    attr_val['collect'] = '%.2f' % line.total_paid or 0.0
                    curr_coll += line.total_paid
                    bal += line.amount - line.concession_amount
                    t_lst.append(attr_val)
                    term_line_empty += 1
                refund += line.refund_amount
            if term_line_empty < aa :
                while (term_line_empty < aa):
                    attr_val = {}
                    attr_val['charge'] = ('%.2f' % 0.0)
                    attr_val['disc_type'] = ''
                    attr_val['disc'] = ('%.2f' % 0.0)
                    attr_val['collect'] = ('%.2f' % 0.0)
                    t_lst.append(attr_val)
                    term_line_empty += 1

            for t in t_lst:
                disc_lst.append(t['charge'])
                disc_lst.append(t['disc_type'])
                disc_lst.append(t['disc'])
                collect_lst.append(t['collect'])
            balance = bal - curr_coll
            ws.append([status,inactive_type, '', '', '', student_number, rte_status, school, school_desc, student_name, grade,
                       ('%.2f' % reg_fee), ('%.2f' % reg_fee_paid),('%.2f' % adm_fee), ('%.2f' % adm_fee_paid), reg_adm_disc_type,
                       ('%.2f' % reg_adm_disc)] + disc_lst + collect_lst + [('%.2f' % refund),('%.2f' % 0.0), ('%.2f' % 0.0), ('%.2f' % balance),('%.2f' % curr_coll), ('%.2f' % 0.0)])
            ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="left")
            ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="left")
            ws['F' + str(t_count + 1)].alignment = Alignment(horizontal="left")
            ws['G' + str(t_count + 1)].alignment = Alignment(horizontal="left")
            ws['H' + str(t_count + 1)].alignment = Alignment(horizontal="left")
            ws['I' + str(t_count + 1)].alignment = Alignment(horizontal="left")
            ws['J' + str(t_count + 1)].alignment = Alignment(horizontal="left")
            alphabets = ['C', 'D', 'E','K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ']
            col = 0
            cnt = 23 + len(disc_list) + len(clt_list)
            for al in alphabets:
                col += 1
                if col <= cnt:
                    ws[al + str(t_count + 1)].alignment = Alignment(horizontal="right")
            border = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S','T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ']
            bor = 0
            thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
            for al in border:
                bor += 1
                if bor <= cnt:
                    ws.cell(row=(t_count + 1), column=(bor)).border = thin_border
            t_count += 1

        #Company Details
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=15, bold=True)
        header1 = NamedStyle(name="header1", font=ft1)
        ws['A1'].style = header1
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:K1')
        ft2 = Font(size=11, bold=True)
        header2 = NamedStyle(name="header2", font=ft2)
        ws['A2'].style = header2
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:K2')
        ws['A3'].style = header2
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:K3')
        ws['A4'].style = header2
        ws['A4'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A4:K4')
        ws['A5'].style = header2
        ws['A5'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A5:K5')
        ws.merge_cells('A6:K6')
        ft4 = Font(size=12, bold=True)
        header3 = NamedStyle(name="header3", font=ft4)
        ws['A7'].style = header3
        ws['A7'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A7:K7')
        ws.merge_cells('A8:K8')
        alphabets = ['A','B','C','D','E','F','G','H','I','J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T','U', 'V', 'W', 'X', 'Y', 'Z','AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ']
        col = 0
        cnt = 23 + len(disc_list) + len(clt_list)
        for al in alphabets:
            col += 1
            if col <= cnt:
                ws[al + str(9)].style = header3
                ws[al + str(9)].alignment = Alignment(horizontal="center")
                thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
                ws.cell(row=9, column=(col)).border = thin_border
                ws.column_dimensions[al].width = 25

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data

    @api.multi
    def generate_excel_report(self):
        # if not self.get_schoolwise_data():
        #     raise ValidationError(_("No record found..!"))
        data = base64.encodestring(self.generate_xl())
        attach_vals = {
            'name': '%s.xls' % ('Branch wise Consolidated Report'),
            'datas': data,
            'datas_fname': '%s.xls' % ('Branch wise Consolidated Report'),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }

SchoolwiseConsolidatedWizard()


