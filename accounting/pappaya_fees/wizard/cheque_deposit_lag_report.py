from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
from odoo.tools.misc import DEFAULT_SERVER_DATE_FORMAT
import base64
from io import BytesIO
from datetime import datetime, date
import re


class PappayaChequeDepositLag(models.TransientModel):
    _name = "pappaya.cheque.deposit.lag"

    branch_ids = fields.Many2many('operating.unit', string='Branch')
    academic_year_id = fields.Many2one('academic.year', string="Academic Year", default=lambda self: self.env['academic.year'].search([('is_active', '=', True)]))
    from_date = fields.Date(string="From Date")
    to_date = fields.Date(string="To Date")
    cheque_deposit_lag_line_ids = fields.One2many('pappaya.cheque.deposit.lag.line', 'cheque_deposit_lag_id',)

    @api.constrains('from_date', 'to_date')
    def _check_dates(self):
        if self.from_date and self.to_date and self.from_date > self.to_date:
            raise ValidationError('End Date must be greater than Start Date')

    @api.onchange('branch_ids', 'academic_year_id', 'from_date', 'to_date')
    def onchange_of_get_data(self):
        domain, data_list = [], []
        if self.branch_ids and self.academic_year_id:
            if self.branch_ids:
                domain.append(('bank_clearance_id.school_id', 'in', self.branch_ids.ids))
                domain.append(('payment_mode_id.is_cheque', '=', True))
                domain.append(('receipt_id.academic_year_id', '=', self.academic_year_id.id))
            if self.from_date:
                from_date = self.from_date + ' 00:00:01'
                domain.append(('create_date', '>=', from_date))
            if self.to_date:
                to_date = self.to_date + ' 23:59:59'
                domain.append(('create_date', '<=', to_date))

            if self.from_date and self.to_date and self.from_date > self.to_date:
                raise ValidationError('End Date must be greater than Start Date')

            for fees_payment in self.env['payment.status.line'].search(domain):
                vals = {}
                if fees_payment.reason:
                    vals['admission_no'] = fees_payment.receipt_id.admission_number
                    vals['student_name'] = fees_payment.receipt_id.enquiry_id.student_full_name
                    vals['receipt_date'] = fees_payment.receipt_date
                    vals['cheque_date'] = fees_payment.cheque_date
                    vals['cheque_dd'] = fees_payment.cheque_dd
                    vals['clearance_date'] = fees_payment.clear_date if fees_payment.state == 'cleared' else False
                    vals['amount'] = fees_payment.total
                    vals['reason'] = fees_payment.reason
                    vals['reject_reason'] = fees_payment.remarks
                    vals['status'] = fees_payment.state
                    data_list.append((0, 0,vals))
            if not data_list:
                self.cheque_deposit_lag_line_ids = None
                return {
                    'warning': {'title': _('Alert'), 'message': _('No record found..!')},
                }
            else:
                self.cheque_deposit_lag_line_ids = data_list

    @api.multi
    def get_deposit_lag_data(self):
        domain, collection_list, data_list = [], [], []
        if self.branch_ids:
            domain.append(('bank_clearance_id.school_id', 'in', self.branch_ids.ids))
            domain.append(('payment_mode_id.is_cheque', '=', True))
        if self.academic_year_id:
            domain.append(('receipt_id.academic_year_id', '=', self.academic_year_id.id))
        if self.from_date:
            from_date = self.from_date + ' 00:00:01'
            domain.append(('create_date', '>=', from_date))
        if self.to_date:
            to_date = self.to_date + ' 23:59:59'
            domain.append(('create_date', '<=', to_date))

        count = 0
        for fees_payment in self.env['payment.status.line'].search(domain):
            vals = {}
            if fees_payment.reason:
                count += 1
                vals['sl_no'] = count
                vals['admission_no'] = fees_payment.receipt_id.admission_number if fees_payment.receipt_id.admission_number else '--'
                vals['stud_name'] = str(fees_payment.receipt_id.enquiry_id.student_full_name) or '--'
                vals['receipt_date'] = (datetime.strptime(str(fees_payment.receipt_date), DEFAULT_SERVER_DATE_FORMAT).strftime("%d/%m/%Y")) if fees_payment.receipt_date else ''
                vals['cheque_date'] = (datetime.strptime(str(fees_payment.cheque_date), DEFAULT_SERVER_DATE_FORMAT).strftime("%d/%m/%Y")) if fees_payment.cheque_date else ''
                vals['cheque_dd'] = fees_payment.cheque_dd or ''
                vals['clearance_date'] = (datetime.strptime(str(fees_payment.clear_date), DEFAULT_SERVER_DATE_FORMAT).strftime("%d/%m/%Y")) if fees_payment.state=='cleared' else ''
                vals['amount'] = fees_payment.total or 0.0
                vals['reason'] = fees_payment.reason or ''
                vals['reject_reason'] = fees_payment.remarks or ''
                vals['status'] = str(fees_payment.state).title() or ''
                data_list.append(vals)
        return data_list

    @api.multi
    def generate_xl(self):
        if not self.get_deposit_lag_data():
            raise ValidationError(_("No record found..!"))
        wb = Workbook()
        ws = wb.active
        ws.title = "Cheque or DD Deposit Lag"
        ws.append([(self.env.user.company_id.name if self.env.user.company_id.name else '')])
        ws.append([(self.env.user.company_id.tem_street if self.env.user.company_id.tem_street else '') + ', ' +
                   (self.env.user.company_id.tem_street2 if self.env.user.company_id.tem_street2 else '') + ', ' +
                   (self.env.user.company_id.tem_city_id.name if self.env.user.company_id.tem_city_id.name else '')+', '+
                   (self.env.user.company_id.tem_state_id.name if self.env.user.company_id.tem_state_id.name else '')+' - '+
                   str(self.env.user.company_id.tem_zip if self.env.user.company_id.tem_zip else '')])
        ws.append(['Tel: ' + (self.env.user.company_id.phone if self.env.user.company_id.phone else '') + ', ' +
                   'Fax: ' + (self.env.user.company_id.fax_id if self.env.user.company_id.fax_id else '') + ', ' +
                   'Email: ' + (self.env.user.company_id.email if self.env.user.company_id.email else '')])
        ws.append([self.env.user.company_id.website if self.env.user.company_id.website else ''])
        ws.append([])
        ws.append(['CHEQUE/DD DEPOSIT LAG REPORT'])
        ws.append(['  For the period of ' + (
            datetime.strptime(str(self.from_date), DEFAULT_SERVER_DATE_FORMAT).strftime("%d/%m/%Y")) + ' to ' +
                   (datetime.strptime(str(self.to_date), DEFAULT_SERVER_DATE_FORMAT).strftime("%d/%m/%Y"))])
        ws.append(['ACADEMIC YEAR: '+ str(self.academic_year_id.name)])
        ws.append([])
        ws.append(['SL NO.','RECEIPT DATE', 'ADMISSION NO.', 'STUDENT NAME', 'REFERENCE NO.', 'CHEQUE/DD DATE',
                   'PAID AMOUNT', 'DEPOSITED LATE REASON', 'CLEARANCE DATE', 'REJECT REASON', 'STATUS'])
        t_count = 10
        # Fetch data
        for xl_list in self.get_deposit_lag_data():
            ws.append([xl_list['sl_no'], xl_list['receipt_date'], xl_list['admission_no'], xl_list['stud_name'],
                        xl_list['cheque_dd'], xl_list['cheque_date'], xl_list['amount'], xl_list['reason'],
                       xl_list['clearance_date'], xl_list['reject_reason'], xl_list['status']])
            ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['C' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['D' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['E' + str(t_count + 1)].alignment = Alignment(horizontal="left")
            ws['F' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['G' + str(t_count + 1)].number_format = '0.00'
            ws['H' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['I' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['J' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['K' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['L' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['M' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            t_count += 1
        # Company Details
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=15, bold=True)
        header1 = NamedStyle(name="header1", font=ft1)
        ws['A1'].style = header1
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:I1')
        ft2 = Font(size=11)
        header2 = NamedStyle(name="header2", font=ft2)
        ws['A2'].style = header2
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:I2')
        ws['A3'].style = header2
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:I3')
        ws['A4'].style = header2
        ws['A4'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A4:I4')
        ft3 = Font(size=10)
        header3 = NamedStyle(name="header3", font=ft3)
        ws['A5'].style = header3
        ws['A5'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A5:I5')
        # Student Payment Status
        ft4 = Font(size=12, bold=True)
        header4 = NamedStyle(name="header4", font=ft4)
        ws['A6'].style = header4
        ws['A6'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A6:I6')
        ws['A7'].style = header3
        ws['A7'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A7:I7')
        # For the period
        ft5 = Font(size=11)
        header5 = NamedStyle(name="header5", font=ft5)
        ws['A8'].style = header3
        ws['A8'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A8:I8')
        ws['A9'].style = header4
        ws['A9'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A9:I9')

        # Header ALignment
        for cell in ws["10"]:
            cell.style = header4
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 18
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data

    @api.multi
    def generate_excel_report(self):
        data = base64.encodestring(self.generate_xl())
        attach_vals = {
            'name': '%s.xls' %('Cheque or DD Deposit Lag'),
            'datas': data,
            'datas_fname': '%s.xls' %('Cheque or DD Deposit Lag'),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }


PappayaChequeDepositLag()


class PappayaChequeDepositLag(models.TransientModel):
    _name = "pappaya.cheque.deposit.lag.line"

    cheque_deposit_lag_id = fields.Many2one('pappaya.cheque.deposit.lag')
    receipt_date = fields.Date('Receipt Date')
    admission_no = fields.Char('Admission No.',size=150)
    student_name = fields.Char('Student Name',size=150)
    cheque_dd = fields.Char('Ref. No',size=30)
    cheque_date = fields.Date('Cheque Date')
    reject_reason = fields.Text(string='Reject Reason',size=300)
    reason = fields.Text(string='Deposited Late Reason',size=300)
    clearance_date = fields.Date('Clearance Date')
    amount = fields.Float('Paid Amount')
    status = fields.Selection([('cleared', 'Cleared'), ('uncleared', 'Uncleared'), ('rejected', 'Rejected')], string='Status')
