from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
import time
from datetime import datetime
from odoo.tools.misc import DEFAULT_SERVER_DATE_FORMAT
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import base64
from io import BytesIO
import logging
_logger = logging.getLogger(__name__)

class BankLedgerWizard(models.TransientModel):
    _name = "bank.ledger.wizard"

    school_id = fields.Many2one('operating.unit',string='Branch')
    academic_year_id = fields.Many2one('academic.year',string='Academic Year', default=lambda self: self.env['academic.year'].search([('is_active', '=', True)]))
    bank_account_id = fields.Many2one('res.bank', string='Bank Name')
    from_date = fields.Date(string='From Date')
    to_date = fields.Date(string='To Date')

    @api.constrains('from_date', 'to_date')
    def check_date(self):
        if self.from_date and self.to_date and self.from_date > self.to_date:
            raise ValidationError(_('To Date should be greater than the From Date!'))
        if self.from_date and self.from_date > time.strftime('%Y-%m-%d'):
            raise ValidationError('From Date is in the future!')
        if self.to_date and self.to_date > time.strftime('%Y-%m-%d'):
            raise ValidationError('To Date is in the future!')

    @api.onchange('school_id')
    def onchange_school(self):
        if self.school_id:
            self.bank_account_id = None
            self.from_date = None
            self.to_date = None

    @api.onchange('academic_year_id')
    def onchange_academic_year(self):
        if self.academic_year_id:
            self.bank_account_id = None
            self.from_date = None
            self.to_date = None

    @api.onchange('from_date')
    def onchange_date(self):
        if self.from_date:
            self.to_date = None

    @api.multi
    def generate_xl(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "CASH BOOK"
        ws.append([(self.school_id.name if self.school_id.name else '')])
        ws.append([(self.school_id.street if self.school_id.street else '') + ', ' + (self.school_id.street2 if self.school_id.street2 else '') + ', ' + (self.school_id.city if self.school_id.city else '')])
        ws.append(['Tel: ' + (self.school_id.mobile if self.school_id.mobile else '') + ', ' + 'Fax: ' + (self.school_id.fax_id if self.school_id.fax_id else '') + ', ' + 'Email: ' + (self.school_id.email if self.school_id.email else '')])
        ws.append([self.school_id.website if self.school_id.website else ''])
        ws.append([])
        ws.append([])
        ws.append(['CASH BOOK'])
        ws.append(['' + (datetime.strptime(str(self.from_date),DEFAULT_SERVER_DATE_FORMAT).strftime('%d/%b/%Y') + ' - ' + datetime.strptime(str(self.to_date),DEFAULT_SERVER_DATE_FORMAT).strftime('%d/%b/%Y'))])
        ws.append([])
        ws.append(['Date', 'Opening Balance', 'Cash Collection', 'Bank Deposit', 'Closing Balance', 'Uncleared Deposit'])
        t_count = 10
        deposit_obj = self.env['bank.deposit'].sudo().search([('school_id','=',self.school_id.id),('academic_year_id', '=', self.academic_year_id.id),('c_bank_id', '=', self.bank_account_id.id),
                                                              ('deposit_date', '>=', self.from_date),('deposit_date', '<=', self.to_date),('state','in',['requested','approved'])])
        for obj in deposit_obj:
            unclr,clr = 0.0,0.0
            date = datetime.strptime(str(obj.deposit_date),DEFAULT_SERVER_DATE_FORMAT).strftime('%d-%b-%Y')
            if obj.state == 'requested':
                unclr += obj.c_amt_deposit
            if obj.state == 'approved':
                clr += obj.c_amt_deposit
            ws.append([date,('%.2f' % obj.opening_bal),('%.2f' % obj.total_cash_amt),('%.2f' % clr),('%.2f' % obj.closing_bal),('%.2f' % unclr)])
            ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['C' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['D' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['E' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['F' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
            ws.cell(row=t_count + 1, column=1).border = thin_border
            ws.cell(row=t_count + 1, column=2).border = thin_border
            ws.cell(row=t_count + 1, column=3).border = thin_border
            ws.cell(row=t_count + 1, column=4).border = thin_border
            ws.cell(row=t_count + 1, column=5).border = thin_border
            ws.cell(row=t_count + 1, column=6).border = thin_border
            t_count += 1

        # Company Details
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=15, bold=True)
        header1 = NamedStyle(name="header1", font=ft1)
        ft2 = Font(size=10)
        header2 = NamedStyle(name="header2", font=ft2)
        ft3 = Font(size=11, bold=True)
        header3 = NamedStyle(name="header3", font=ft3)
        ws['A1'].style = header1
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:F1')
        ws['A2'].style = header2
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:F2')
        ws['A3'].style = header2
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:F3')
        ws['A4'].style = header2
        ws['A4'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A4:F4')
        ws['A5'].style = header2
        ws['A5'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A5:F5')
        ws['A6'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A6:F6')
        # styles for row7
        ws['A7'].style = header3
        ws['A7'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A7:F7')
        # Styles for row 8,9
        ws['A8'].style = header2
        ws['A8'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A8:F8')
        ws.merge_cells('A9:F9')
        # Styles for row 10
        ws['A10'].style = header3
        ws['A10'].alignment = Alignment(horizontal="center")
        ws['B10'].style = header3
        ws['B10'].alignment = Alignment(horizontal="center")
        ws['C10'].style = header3
        ws['C10'].alignment = Alignment(horizontal="center")
        ws['D10'].style = header3
        ws['D10'].alignment = Alignment(horizontal="center")
        ws['E10'].style = header3
        ws['E10'].alignment = Alignment(horizontal="center")
        ws['F10'].style = header3
        ws['F10'].alignment = Alignment(horizontal="center")
        # Borders
        thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
        ws.cell(row=10, column=1).border = thin_border
        ws.cell(row=10, column=2).border = thin_border
        ws.cell(row=10, column=3).border = thin_border
        ws.cell(row=10, column=4).border = thin_border
        ws.cell(row=10, column=5).border = thin_border
        ws.cell(row=10, column=6).border = thin_border
        # Width style
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data

    @api.multi
    def generate_excel_report(self):
        data = base64.encodestring(self.generate_xl())
        attach_vals = {
            'name': '%s.xls' % ('Cash Book'),
            'datas': data,
            'datas_fname': '%s.xls' % ('Cash Book'),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }