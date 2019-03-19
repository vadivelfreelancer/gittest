from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from datetime import datetime, date
from odoo.tools.misc import DEFAULT_SERVER_DATE_FORMAT
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import base64
import os
from io import BytesIO
import logging
_logger = logging.getLogger(__name__)

class FeeConcessionWizard(models.TransientModel):
    _name = "fee.concession.wizard"

    society_ids =fields.Many2many('operating.unit','society_concession_strength_rel','company_id','society_id',string= 'Society')
    school_ids = fields.Many2many('operating.unit','school_concession_rel','company_id','school_id', string='Branch')
    academic_year_ids = fields.Many2many('academic.year','academic_year_concession_rel','school_id', 'academic_year_id', string='Academic Year')
    student_ids = fields.Many2many('pappaya.student','student_concession_rel','school_id', 'student_id', string='Student')

    @api.onchange('society_ids')
    def onchange_society_ids(self):
        if self.society_ids:
            self.school_ids = []
            return {'domain': {'school_ids': [('type', '=', 'school'),('parent_id', 'in', self.society_ids.ids)]}}
    
    @api.onchange('school_ids')
    def onchange_school_id(self):
        if self.school_ids:
            year_ids = self.academic_ids()
            return {'domain': {'academic_year_ids': [('id', 'in', year_ids)]}}
            
    @api.onchange('academic_year_ids')
    def onchange_student_id(self):
        if self.academic_year_ids:
            return {'domain': {'student_ids': [('school_id', 'in', self.school_ids.ids),('society_id', 'in', self.society_ids.ids)]}}        
            
    @api.multi
    def academic_ids(self):
        date_list = []
        year_ids = []
        if self.school_ids:
            max_start_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="start_date desc",limit=1).start_date
            min_start_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="start_date",limit=1).start_date
            max_end_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="end_date desc",limit=1).end_date
            min_end_date = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)],order="end_date",limit=1).end_date
            date_list.append(max_start_date)
            date_list.append(min_start_date)
            date_list.append(max_end_date)
            date_list.append(min_end_date)
            min_year = (datetime.strptime(min(date_list), DEFAULT_SERVER_DATE_FORMAT)).year
            max_year = (datetime.strptime(max(date_list), DEFAULT_SERVER_DATE_FORMAT)).year 
            academic_year_ids = self.env['academic.year'].sudo().search([('school_id', 'in', self.school_ids.ids)])
            for academic_year in  academic_year_ids:
                if (datetime.strptime(academic_year.start_date, DEFAULT_SERVER_DATE_FORMAT)).year >= min_year or (datetime.strptime(academic_year.start_date, DEFAULT_SERVER_DATE_FORMAT)).year <= max_year:
                    year_ids.append(academic_year.id)
            return year_ids

    @api.multi
    def get_school_header(self):
        school_list = []
        if (len(self.school_ids) == 1 and len(self.society_ids) == 1) or (
                len(self.school_ids) == 1 and not self.society_ids):
            vals = {}
            vals['school_id'] = self.school_ids.name
            vals['logo'] = self.school_ids.logo if self.school_ids.logo else ''
            vals['street'] = self.school_ids.street if self.school_ids.street else ''
            vals['street2'] = self.school_ids.street2 if self.school_ids.street2 else ''
            vals['city'] = self.school_ids.city if self.school_ids.city else ''
            vals['zip'] = self.school_ids.zip if self.school_ids.zip else ''
            vals['phone'] = self.school_ids.phone if self.school_ids.phone else ''
            vals['fax'] = self.school_ids.fax_id if self.school_ids.fax_id else ''
            vals['email'] = self.school_ids.email if self.school_ids.email else ''
            vals['website'] = self.school_ids.website if self.school_ids.website else ''
            school_list.append(vals)
        if (len(self.society_ids) == 1 and not self.school_ids):
            vals = {}
            vals['school_id'] = self.society_ids.name
            vals['logo'] = self.society_ids.logo
            vals['street'] = self.society_ids.street if self.society_ids.street else ''
            vals['street2'] = self.society_ids.street2 if self.society_ids.street2 else ''
            vals['city'] = self.society_ids.city if self.society_ids.city else ''
            vals['zip'] = self.society_ids.zip if self.society_ids.zip else ''
            vals['phone'] = self.society_ids.phone if self.society_ids.phone else ''
            vals['fax'] = self.society_ids.fax_id if self.society_ids.fax_id else ''
            vals['email'] = self.society_ids.email if self.society_ids.email else ''
            vals['website'] = self.society_ids.website if self.society_ids.website else ''
            school_list.append(vals)
        if (len(self.society_ids) == 1 and len(self.school_ids) != 1):
            vals = {}
            vals['school_id'] = self.society_ids.name
            vals['logo'] = self.society_ids.logo
            vals['street'] = self.society_ids.street if self.society_ids.street else ''
            vals['street2'] = self.society_ids.street2 if self.society_ids.street2 else ''
            vals['city'] = self.society_ids.city if self.society_ids.city else ''
            vals['zip'] = self.society_ids.zip if self.society_ids.zip else ''
            vals['phone'] = self.society_ids.phone if self.society_ids.phone else ''
            vals['fax'] = self.society_ids.fax_id if self.society_ids.fax_id else ''
            vals['email'] = self.society_ids.email if self.society_ids.email else ''
            vals['website'] = self.society_ids.website if self.society_ids.website else ''
            school_list.append(vals)
        return school_list

    @api.multi
    def get_society_header(self):
        society_list = []
        if not self.society_ids and not self.school_ids:
            soc_list = ''
            obj = self.env['res.company'].sudo().search([('type', '=', 'society')])
            for record in obj:
                soc_list += str(record.name) + ', '
            vals = {}
            vals['society_id'] = soc_list[:-2]
            society_list.append(vals)
        else:
            sc_list = ''
            for record in self.society_ids:
                sc_list += str(record.name) + ', '
            vals = {}
            vals['society_id'] = sc_list[:-2]
            society_list.append(vals)
        return society_list

    @api.multi
    def get_concession_data(self):
        domain, data_list, conc_dict = [], [], {}
        if self.society_ids:
            domain.append(('society_id','in',self.society_ids.ids))
        if self.school_ids:
            domain.append(('school_id','in',self.school_ids.ids))
        if self.academic_year_ids:
            domain.append(('academic_year_id','in',self.academic_year_ids.ids))
        if self.student_ids:
            domain.append(('student_id','in',self.student_ids.ids))
        fee_obj = self.env['pappaya.fees.concession'].sudo().search([('state','=','applied')]+domain)
        for obj in fee_obj:
            for line in obj.concession_head_line:
                if line.concession_applied == True:
                    vals = {}
                    vals['society'] = obj.society_id.name if obj.society_id else ''
                    vals['school'] = obj.school_id.name if obj.school_id else ''
                    vals['academic'] = obj.academic_year_id.name if obj.academic_year_id else ' '
                    vals['stud_id'] = obj.student_id.enrollment_num if obj.student_id.enrollment_num else ''
                    vals['stud_name'] = str(obj.student_id.name) + ' ' + str(obj.student_id.middle_name or '') + ' ' + str(obj.student_id.last_name or '')
                    vals['type'] = obj.concession_type_id.code if obj.concession_type_id.code else ''
                    vals['fee_head'] = line.name if line.name else ''
                    vals['amount'] = ('%.2f' % line.total_amount) if line.total_amount else ('%.2f' %0.00)
                    vals['remarks'] = obj.reason if obj.reason else ''
                    data_list.append(vals)
        return data_list

    @api.multi
    def generate_xl(self):
        wb= Workbook()
        ws= wb.active
        ws.title = "FEE CONCESSION"
        if (len(self.school_ids) == 1 and len(self.society_ids) == 1) or (len(self.school_ids) == 1 and not self.society_ids):
            ws.append([(self.school_ids.name if self.school_ids.name else '')])
            ws.append([(self.school_ids.street if self.school_ids.street else '') + ', ' + (self.school_ids.street2 if self.school_ids.street2 else '') + ', ' + (self.school_ids.city if self.school_ids.city else '')])
            ws.append(['Tel: ' + (self.school_ids.phone if self.school_ids.phone else '') + ', ' + 'Fax: ' + (self.school_ids.fax_id if self.school_ids.fax_id else '') + ', ' + 'Email: ' + (self.school_ids.email if self.school_ids.email else '')])
            ws.append([self.school_ids.website if self.school_ids.website else ''])
            ws.append([])
        if (len(self.society_ids) == 1 and not self.school_ids) or (len(self.society_ids) == 1 and len(self.school_ids) > 1):
            ws.append([(self.society_ids.name if self.society_ids.name else '')])
            ws.append([(self.society_ids.street if self.society_ids.street else '') + ', ' + (self.society_ids.street2 if self.society_ids.street2 else '') + ', ' + (self.society_ids.city if self.society_ids.city else '')])
            ws.append(['Tel: ' + (self.society_ids.phone if self.society_ids.phone else '') + ', ' + 'Fax: ' + (self.society_ids.fax_id if self.society_ids.fax_id else '') + ', ' + 'Email: ' + (self.society_ids.email if self.society_ids.email else '')])
            ws.append([self.society_ids.website if self.society_ids.website else ''])
            ws.append([])
        if (not self.society_ids and not self.school_ids) or (len(self.school_ids) > 1 and not self.society_ids):
            soc_list = ''
            obj = self.env['res.company'].sudo().search([('type', '=', 'society')])
            for record in obj:
                soc_list += str(record.name) + ', '
            ws.append([])
            ws.append([])
            ws.append([soc_list[:-2]])
            ws.append([])
            ws.append([])
        if (len(self.society_ids) > 1) or (len(self.society_ids) > 1 and len(self.school_ids) > 1):
            sc_list = ''
            for record in self.society_ids:
                sc_list += str(record.name) + ', '
            ws.append([])
            ws.append([])
            ws.append([sc_list[:-2]])
            ws.append([])
            ws.append([])
        ws.append([])
        ws.append(['FEE CO NCESSION'])
        ws.append([])
        ws.append(['Society','Branch','Academic Year','Student ID','Student Name','Concession Type','Fee Type','Amount','Remarks'])
        t_count = 9
        # Fetch data
        for xl_list in self.get_concession_data():
            ws.append([xl_list['society'],xl_list['school'],xl_list['academic'],xl_list['stud_id'],xl_list['stud_name'],xl_list['type'],
                       xl_list['fee_head'],xl_list['amount'],xl_list['remarks']])
            ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['C' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['D' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['E' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['F' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['G' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['H' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['I' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['H' + str(t_count + 1)].number_format = '0.0'
            t_count += 1
        #Company Details
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=15, bold=True)
        header1 = NamedStyle(name="header1", font=ft1)
        ws['A1'].style = header1
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:I1')
        ft2 = Font(size=11, bold=True)
        header2 = NamedStyle(name="header2", font=ft2)
        ws['A2'].style = header2
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:I2')
        ws['A3'].style = header2
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:I3')
        ws['A4'].style = header2
        ws['A4'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A4:I4')
        ws['A5'].style = header2
        ws['A5'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A5:I5')
        ws['A6'].style = header2
        ws['A6'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A6:I6')
        # styles for row7
        ft4 = Font(size=12, bold=True)
        header3 = NamedStyle(name="header3", font=ft4)
        ws['A7'].style = header3
        ws['A7'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A7:I7')
        ft6 = Font(size=12)
        header6 = NamedStyle(name="header6", font=ft6)
        ws['A8'].style = header6
        ws['A8'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A8:I8')
        # styles for row9
        ft5 = Font(size=11,bold=True)
        header4 = NamedStyle(name="header4", font=ft5)
        ws['A9'].style = header4
        ws['A9'].alignment = Alignment(horizontal="center")
        ws['B9'].style = header4
        ws['B9'].alignment = Alignment(horizontal="center")
        ws['C9'].style = header4
        ws['C9'].alignment = Alignment(horizontal="center")
        ws['D9'].style = header4
        ws['D9'].alignment = Alignment(horizontal="center")
        ws['E9'].style = header4
        ws['E9'].alignment = Alignment(horizontal="center")
        ws['F9'].style = header4
        ws['F9'].alignment = Alignment(horizontal="center")
        ws['G9'].style = header4
        ws['G9'].alignment = Alignment(horizontal="center")
        ws['H9'].style = header4
        ws['H9'].alignment = Alignment(horizontal="center")
        ws['I9'].style = header4
        ws['I9'].alignment = Alignment(horizontal="center")
        thin_border = Border(top=Side(style='thin'),bottom=Side(style='thin'), right=Side(style='thin'),left=Side(style='thin'))
        ws.cell(row=9, column=1).border = thin_border
        ws.cell(row=9, column=2).border = thin_border
        ws.cell(row=9, column=3).border = thin_border
        ws.cell(row=9, column=4).border = thin_border
        ws.cell(row=9, column=5).border = thin_border
        ws.cell(row=9, column=6).border = thin_border
        ws.cell(row=9, column=7).border = thin_border
        ws.cell(row=9, column=8).border = thin_border
        ws.cell(row=9, column=9).border = thin_border
        # Width style
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18

        # cwd = os.path.abspath(__file__)
        # path = cwd.rsplit('/', 2)
        # img_path = path[0] + '/static/src/img/dis_logo.png'
        # image_path = img_path.replace("pappaya_fees", "pappaya_core")
        # img = openpyxl.drawing.image.Image(image_path)
        # ws.add_image(img, 'A1')
        # ws['A1'].alignment = Alignment(horizontal="center")
        # ws.merge_cells('A1:A5')

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data

    @api.multi
    def generate_pdf_report(self):
        if not self.get_concession_data():
            raise ValidationError(_("No record found..!"))
        if self.get_concession_data():
            return self.env['report'].get_action(self, 'pappaya_fees.template_fee_concession')

    @api.multi
    def generate_excel_report(self):
        if not self.get_concession_data():
            raise ValidationError(_("No record found..!"))
        data = base64.encodestring(self.generate_xl())
        attach_vals = {
            'name': '%s.xls' % ('Fee Concession'),
            'datas': data,
            'datas_fname': '%s.xls' % ('Fee Concession'),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }

FeeConcessionWizard()


