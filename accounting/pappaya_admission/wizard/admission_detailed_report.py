from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side
import base64
from io import BytesIO


class AdmissionDetailedReport(models.TransientModel):
    _name = "admission.detailed.report"

    branch_id = fields.Many2one('operating.unit', string='Branch')
    academic_year_id  = fields.Many2one('academic.year', string="Academic Year")
    course_package_ids = fields.Many2many('pappaya.course.package', string='Course Package')
    stage_ids = fields.Many2many('pappaya.business.stage', string='Status')

    @api.onchange('branch_id')
    def onchange_branch(self):
        domain = []
        if self.branch_id:
            self.academic_year_id = None
            self.course_package_ids = None
            self.stage_ids = None
            for academic in self.branch_id.course_config_ids:
                domain.append(academic.academic_year_id.id)
        return {'domain': {'academic_year_id': [('id', 'in', domain)]}}

    @api.onchange('academic_year_id','branch_id')
    def onchange_course_package(self):
        domain = []
        if self.academic_year_id and self.branch_id:
            for academic in self.branch_id.course_config_ids:
                if academic.academic_year_id.id == self.academic_year_id.id:
                    for course_package in academic.course_package_ids:
                        domain.append(course_package.id)
        return {'domain': {'course_package_ids': [('id', 'in', domain)]}}


    @api.multi
    def get_data(self):
        domain, adm_dict, data_list = [], {}, []
        if self.branch_id:
            domain.append(('branch_id', '=', self.branch_id.id))
        if self.academic_year_id:
            domain.append(('academic_year', '=', self.academic_year_id.id))
        if self.course_package_ids:
            domain.append(('package_id', 'in', self.course_package_ids.ids))
        if self.stage_ids:
            domain.append(('stage_id', 'in', self.stage_ids.ids))
        admission_obj = self.env['pappaya.admission'].search([('cancel','=',False)]+domain)
        for obj in admission_obj:
            if obj.branch_id not in adm_dict:
                adm_dict.update({obj.branch_id: [obj]})
            else:
                adm_dict[obj.branch_id].append(obj)
        for key in adm_dict:
            for obj in admission_obj:
                if key.name == obj.branch_id.name:
                    vals = {}
                    vals['branch_name'] = key.name if key.name else ''
                    vals['admission_no'] = obj.res_no if obj.res_no else ''
                    vals['student_name'] = str(obj.sur_name or '') + ' ' + str(obj.name)
                    vals['course_package_name'] = obj.package_id.name if obj.package_id else ''
                    vals['status_name'] = obj.stage_id.name if obj.stage_id else ''
                    data_list.append(vals)
        return data_list

    @api.multi
    def get_stage_data(self):
        domain, domain1, data_list = [], [], []
        if self.branch_id:
            domain.append(('school_id', '=', self.branch_id.id))
        if self.stage_ids:
            domain.append(('id', 'in', self.stage_ids.ids))
        if self.course_package_ids:
            domain1.append(('package_id', 'in', self.course_package_ids.ids))
        vals = {}
        stage_list = []
        stage_obj = self.env['pappaya.business.stage'].search(domain)
        for stage in stage_obj:
            stage_list.append(stage.name)
        sl_list = []
        for sl in stage_list:
            sl_count = 0
            for adm in self.env['pappaya.admission'].search([('cancel','=',False),('stage_id.name', '=', sl),('branch_id', '=', self.branch_id.id),
                                                             ('academic_year', '=', self.academic_year_id.id)]+domain1):
                if adm:
                    sl_count += 1
            sl_vals = {}
            if sl_count > 0.0:
                sl_list.append(sl_count or 0)
                sl_vals['s_count'] = sl_count or 0
            else:
                sl_list.append(0)
                sl_vals['s_count'] = 0
            vals['stage_count']= sl_list
        data_list.append(vals)
        return data_list

    @api.multi
    def generate_xl(self):
        if not self.get_data():
            raise ValidationError(_("No record found..!"))
        wb = Workbook()
        ws = wb.active
        ws.append([(self.branch_id.name if self.branch_id.name else '')])
        ws.append([(self.branch_id.tem_street if self.branch_id.tem_street else '')])
        ws.append([str(self.branch_id.tem_street2 if self.branch_id.tem_street2 else '') + ', ' +
                   str(self.branch_id.tem_city_id.name if self.branch_id.tem_city_id else '') + ', ' +
                   str(self.branch_id.tem_state_id.name if self.branch_id.tem_state_id else '') + '-' +
                   str(self.branch_id.tem_zip if self.branch_id.tem_zip else '')])
        ws.append(['Phone: ' + str(self.branch_id.phone if self.branch_id.phone else '') + ', ' +
                   'Email: ' + str(self.branch_id.email if self.branch_id.email else '')])
        ws.append([self.branch_id.website if self.branch_id.website else ''])
        ws.append([])
        ws.append(['ADMISSION DETAILED REPORT'])
        ws.append([])
        ws.append(['ACADEMIC YEAR : ' + self.academic_year_id.name])
        ws.append([])
        t_count = 10

        # Stage Count
        domain = []
        if self.branch_id:
            domain.append(('school_id', '=', self.branch_id.id))
        if self.stage_ids:
            domain.append(('id', 'in', self.stage_ids.ids))
        stage_list = []
        stage_obj = self.env['pappaya.business.stage'].search(domain)
        for stage in stage_obj:
            stage_list.append(stage.name)
        ws.append(['BRANCH'] + stage_list)
        t_count += 1
        for st_list in self.get_stage_data():
            ws.append([self.branch_id.name] + st_list['stage_count'])
            ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['C' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['D' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['E' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['F' + str(t_count + 1)].alignment = Alignment(horizontal="center")
        t_count += 1

        # Detailed data
        ws.append([])
        ws.append(['BRANCH NAME', 'RES/ADM NUMBER', 'STUDENT NAME', 'COURSE PACKAGE', 'STATUS'])
        t_count += 2
        for xl_list in self.get_data():
            ws.append([xl_list['branch_name'], xl_list['admission_no'], xl_list['student_name'], xl_list['course_package_name'], xl_list['status_name']])
            ws['A' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['B' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['C' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['D' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['E' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            ws['F' + str(t_count + 1)].alignment = Alignment(horizontal="center")
            t_count += 1
        # Company Style
        ws.row_dimensions[1].height = 24
        ft1 = Font(size=14, bold=True)
        header1 = NamedStyle(name="header1", font=ft1)
        ws['A1'].style = header1
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:F1')
        ft2 = Font(size=10)
        header2 = NamedStyle(name="header2", font=ft2)
        ws['A2'].style = header2
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:F2')
        ws['A3'].style = header2
        ws['A3'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A3:F3')
        ws['A4'].style = header2
        ws['A4'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A4:F4')
        ws['A5'].style = header2
        ws['A5'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A5:F5')
        ws['A6'].style = header2
        ws['A6'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A6:F6')
        # Report name style
        ft4 = Font(size=12, bold=True)
        header3 = NamedStyle(name="header3", font=ft4)
        ws['A7'].style = header3
        ws['A7'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A7:F7')
        ws['A8'].style = header3
        ws['A8'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A8:F8')
        # Academic Year Style
        ws['A9'].style = header2
        ws['A9'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A9:F9')
        ws['A10'].style = header2
        ws['A10'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A10:F10')
        # Heading Style
        ws['A11'].style = header3
        ws['A11'].alignment = Alignment(horizontal="center")
        ws['B11'].style = header3
        ws['B11'].alignment = Alignment(horizontal="center")
        ws['C11'].style = header3
        ws['C11'].alignment = Alignment(horizontal="center")
        ws['D11'].style = header3
        ws['D11'].alignment = Alignment(horizontal="center")
        ws['E11'].style = header3
        ws['E11'].alignment = Alignment(horizontal="center")
        ws['F11'].style = header3
        ws['F11'].alignment = Alignment(horizontal="center")
        ws['A14'].style = header3
        ws['A14'].alignment = Alignment(horizontal="center")
        ws['B14'].style = header3
        ws['B14'].alignment = Alignment(horizontal="center")
        ws['C14'].style = header3
        ws['C14'].alignment = Alignment(horizontal="center")
        ws['D14'].style = header3
        ws['D14'].alignment = Alignment(horizontal="center")
        ws['E14'].style = header3
        ws['E14'].alignment = Alignment(horizontal="center")
        ws['F14'].style = header3
        ws['F14'].alignment = Alignment(horizontal="center")
        # Border
        thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin'))
        # ws.cell(row=11, column=1).border = thin_border
        # ws.cell(row=11, column=2).border = thin_border
        # ws.cell(row=11, column=3).border = thin_border
        # ws.cell(row=11, column=4).border = thin_border
        # ws.cell(row=11, column=5).border = thin_border
        # ws.cell(row=14, column=1).border = thin_border
        # ws.cell(row=14, column=2).border = thin_border
        # ws.cell(row=14, column=3).border = thin_border
        # ws.cell(row=14, column=4).border = thin_border
        # ws.cell(row=14, column=5).border = thin_border
        # Width style
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        return data

    @api.multi
    def generate_excel(self):
        data = base64.encodestring(self.generate_xl())
        attach_vals = {
            'name': '%s.xls' % ('Admission Detailed Report'),
            'datas': data,
            'datas_fname': '%s.xls' % ('Admission Detailed Report'),
        }
        doc_id = self.env['ir.attachment'].create(attach_vals)
        return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/%s?download=true' % (doc_id.id),
            'target': 'self',
        }





